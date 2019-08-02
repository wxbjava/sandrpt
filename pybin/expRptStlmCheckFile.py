#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#����ϵͳ���˱���
#�ɴ�������������,����ָ����־����

import cx_Oracle
import sys
from openpyxl.workbook import Workbook
import os
from utl.common import *
from utl.gldict import *
from math import fabs

#���ȫ����
i = 0

#��������ͳ�ƽṹ
class payTxnCount():
    def __init__(self):
        self.count = 0
        self.transAmt = 0
        self.prodAmt = 0
        self.costAmt = 0
        self.rtnAmt = 0

def getPayCountData(payTypeDList, key):
    try:
        result = payTypeDList[key]
    except KeyError:
        result = payTxnCount()
    return result

#ͨ����ֵ����һ�����ܷ���
def getAgentIncome(db, key_rsp):
    cursor = db.cursor()
    sql = "select profits_amt from tbl_ins_profits_txn_dtl where key_rsp = '%s'" % key_rsp
    cursor.execute(sql)
    x = cursor.fetchone()
    if x is not None:
        amt = toNumberFmt(x[0])
    else:
        amt = 0.0
    cursor.close()
    return amt

#ͨ����ֵ����ͨ�����������Ϣ
def getChnlTxn(db, chnlId, chlTxnKey):
    cursor = db.cursor()
    sql = "select MCHT_CD, TERM_CD, REAL_TRANS_AMT, ISS_FEE, SWT_FEE, PROD_FEE  " \
          "from TBL_STLM_TXN_BILL_DTL where CHNL_ID = '%s' and TXN_KEY = '%s'" % (chnlId, chlTxnKey)
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    return x[0], x[1], toNumberFmt(x[2]), toNumberFmt(x[3]), toNumberFmt(x[4]), toNumberFmt(x[5])

#ͨ����ֵ���ҽ�����Ϣ,�����̻���,�ն˺�,���׽��
def getOwnTxn(db, keyRsp) :
    cursor = db.cursor()
    sql = "select card_accp_id, CARD_ACCP_TERM_ID, trans_amt  " \
          "from tbl_acq_txn_log where key_rsp = '%s'" %  keyRsp
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    return x[0], x[1], toNumberFmt(int(x[2]))

#��������˽���
def handleTxn01Rpt(db, ws, stlm_date):
    handleTxn01RptHead(ws)
    handleTxn01RptBody(db, ws, stlm_date)

def handleTxn01RptHead(ws):
    #��������
    ws.cell(row=1, column=8).value = '�յ�����������˱���'
    #����ͷ
    ws.cell(row=2, column=1).value = 'ϵͳ��������'
    ws.cell(row=2, column=2).value = '��Ŀ��ʶ'
    ws.cell(row=2, column=3).value = '���ױ���'
    ws.cell(row=2, column=4).value = '���׽��'
    ws.cell(row=2, column=5).value = '�̻�������'
    ws.cell(row=2, column=6).value = '���������'
    ws.cell(row=2, column=7).value = '��������ת�ӷ�'
    ws.cell(row=2, column=8).value = 'Ʒ�Ʒ����'
    ws.cell(row=2, column=9).value = '�ܳɱ�'
    ws.cell(row=2, column=10).value = '��������'
    ws.cell(row=2, column=11).value = '�ʽ����㾻��'
    ws.cell(row=2, column=12).value = '�յ�����'
    ws.cell(row=2, column=13).value = '�ܲ�����'
    ws.cell(row=2, column=14).value = '�ֹ�˾����'
    ws.cell(row=2, column=15).value = '����������'


def handleTxn01RptBody(db, ws, stlm_date):
    count,allCount = 0, 0
    transAmt,allTransAmt = 0.0, 0.0
    issAmt,allIssAmt = 0.0, 0.0
    swtAmt,allSwtAmt = 0.0, 0.0
    prodAmt,allProdAmt = 0.0, 0.0
    errAmt,allErrAmt = 0.0, 0.0
    mchtFee,allMchtFee = 0.0, 0.0
    companyIncome,allCompanyIncome = 0.0, 0.0
    agentIncome,allAgentIncome = 0.0, 0.0
    old_itemId = ''
    itemId = ''
    #���Ҷ��˳ɹ��Ľ���
    global i
    i = 3
    sql = "select ITEM_ID, key_rsp, REAL_TRANS_AMT, mcht_fee, " \
          "iss_fee, swt_fee, prod_fee, err_fee from TBL_STLM_TXN_BILL_DTL " \
          "where host_date = '%s' and txn_num ='1011' and check_sta ='1' order by ITEM_ID " % (stlm_date)
    cursor = db.cursor()
    cursor.execute(sql)
    for ltTxn in cursor:
        if count == 0:
            old_itemId = ltTxn[0]
        itemId = ltTxn[0]
        if (old_itemId != itemId):
            #��Ŀid��ͬ
            tailTxn01RptBody(ws,i,stlm_date,old_itemId,count,transAmt,issAmt,swtAmt,
                             prodAmt,errAmt,mchtFee,companyIncome,agentIncome)
            i = i + 1
            #��ʼ��
            old_itemId = itemId
            count = 0
            transAmt = 0.0
            issAmt = 0.0
            swtAmt = 0.0
            prodAmt = 0.0
            errAmt = 0.0
            mchtFee = 0.0
            companyIncome = 0.0
            agentIncome = 0.0

        #���Ҵ����̷���
        key_rsp = ltTxn[1]
        amtTmp = getAgentIncome(db, key_rsp)
        agentIncome = toNumberFmt(agentIncome + amtTmp)
        allAgentIncome = toNumberFmt(allAgentIncome + amtTmp)
        count = count + 1
        allCount = allCount + 1
        transAmt = toNumberFmt(transAmt + ltTxn[2])
        allTransAmt = toNumberFmt(allTransAmt + ltTxn[2])
        mchtFee = toNumberFmt(mchtFee + ltTxn[3])
        allMchtFee = toNumberFmt(allMchtFee + ltTxn[3])
        issAmt = toNumberFmt(issAmt + ltTxn[4])
        allIssAmt = toNumberFmt(allIssAmt + ltTxn[4])
        swtAmt = toNumberFmt(swtAmt + ltTxn[5])
        allSwtAmt = toNumberFmt(allSwtAmt + ltTxn[5])
        prodAmt = toNumberFmt(prodAmt + ltTxn[6])
        allProdAmt = toNumberFmt(allProdAmt + ltTxn[6])
        errAmt = toNumberFmt(errAmt + ltTxn[7])
        allErrAmt = toNumberFmt(allErrAmt + ltTxn[7])
    if itemId != '':
        tailTxn01RptBody(ws, i, stlm_date, old_itemId, count, transAmt, issAmt, swtAmt,
                         prodAmt, errAmt, mchtFee, companyIncome, agentIncome)
        i = i + 1

    tailTxn01RptTail(ws, i, allCount, allTransAmt, allIssAmt, allSwtAmt,
                     allProdAmt, allErrAmt, allMchtFee, allCompanyIncome, allAgentIncome)
    i = i + 1

    cursor.close()

def tailTxn01RptBody(ws,i,stlmDate,itemId,count,transAmt,issAmt,swtAmt,
                    prodAmt,errAmt,mchtFee,companyIncome,agentIncome) :
    ws.cell(row=i, column=1).value = stlmDate
    ws.cell(row=i, column=2).value = getItemName(itemId)
    ws.cell(row=i, column=3).value = count
    ws.cell(row=i, column=4).value = transAmt
    ws.cell(row=i, column=5).value = mchtFee
    ws.cell(row=i, column=6).value = issAmt
    ws.cell(row=i, column=7).value = swtAmt
    ws.cell(row=i, column=8).value = prodAmt
    ws.cell(row=i, column=9).value = toNumberFmt(issAmt + swtAmt + prodAmt)
    ws.cell(row=i, column=10).value = errAmt
    ws.cell(row=i, column=11).value = toNumberFmt(transAmt - (issAmt + swtAmt + prodAmt))
    ws.cell(row=i, column=12).value = toNumberFmt(mchtFee - (issAmt + swtAmt + prodAmt))
    ws.cell(row=i, column=13).value = toNumberFmt(mchtFee - (issAmt + swtAmt + prodAmt) - companyIncome - agentIncome)
    ws.cell(row=i, column=14).value = companyIncome
    ws.cell(row=i, column=15).value = agentIncome

def tailTxn01RptTail(ws,i,count,transAmt,issAmt,swtAmt,
                    prodAmt,errAmt,mchtFee,companyIncome,agentIncome) :
    ws.cell(row=i, column=2).value = 'С��'
    ws.cell(row=i, column=3).value = count
    ws.cell(row=i, column=4).value = transAmt
    ws.cell(row=i, column=5).value = mchtFee
    ws.cell(row=i, column=6).value = issAmt
    ws.cell(row=i, column=7).value = swtAmt
    ws.cell(row=i, column=8).value = prodAmt
    ws.cell(row=i, column=9).value = toNumberFmt(issAmt + swtAmt + prodAmt)
    ws.cell(row=i, column=10).value = errAmt
    ws.cell(row=i, column=11).value = toNumberFmt(transAmt - (issAmt + swtAmt + prodAmt))
    ws.cell(row=i, column=12).value = toNumberFmt(mchtFee - (issAmt + swtAmt + prodAmt))
    ws.cell(row=i, column=13).value = toNumberFmt(mchtFee - (issAmt + swtAmt + prodAmt) - companyIncome - agentIncome)
    ws.cell(row=i, column=14).value = companyIncome
    ws.cell(row=i, column=15).value = agentIncome


#��������˽���
def handleTxn02Rpt(db, ws, stlm_date):
    handleTxn02RptHead(ws)
    handleTxn02RptBody(db, ws, stlm_date)

def handleTxn02RptHead(ws):
    global i
    i = i + 2
    #����ͷ
    ws.cell(row=i, column=1).value = '��������'
    ws.cell(row=i, column=2).value = 'S0/T1'
    ws.cell(row=i, column=3).value = '��������'
    ws.cell(row=i, column=4).value = '�������'
    ws.cell(row=i, column=5).value = 'Ʒ�Ʒ����'
    ws.cell(row=i, column=6).value = '�����ɱ�'
    ws.cell(row=i, column=7).value = '�ʽ����㾻��'
    ws.cell(row=i, column=8).value = '�����˵����'
    i = i + 1


def handleTxn02RptBody(db, ws, stlm_date):
    #���Ҷ��˳ɹ��Ľ���
    global i
    sql = "select pay_type, txn_num, count(*), sum(REAL_TRANS_AMT), sum(prod_fee), sum(ISS_FEE) " \
          " from TBL_STLM_TXN_BILL_DTL " \
          "where host_date = '%s' and txn_num in ('1801','9164') and check_sta ='1' and chnl_id ='A002' group by  pay_type, txn_num" % (stlm_date)
    cursor = db.cursor()
    cursor.execute(sql)
    payTypeList = {}
    for ltTxn in cursor:
        data = getPayCountData(payTypeList, ltTxn[0])
        if ltTxn[1] == '1801':
            data.transAmt = ltTxn[3]
            data.count = ltTxn[2]
            data.prodAmt = ltTxn[4]
            data.costAmt = ltTxn[5]
        elif ltTxn[1] == '9164':
            data.rtnAmt = ltTxn[3]
        payTypeList[ltTxn[0]] = data

    allCount = 0
    allTransAmt = 0
    allProdAmt = 0
    allCostAmt = 0
    allRtnAmt = 0

    for payType in payTypeList:
        allCount = allCount + payTypeList[payType].count
        allTransAmt = allTransAmt + payTypeList[payType].transAmt
        allProdAmt = allProdAmt + payTypeList[payType].prodAmt
        allCostAmt = allCostAmt + payTypeList[payType].costAmt
        allRtnAmt = allRtnAmt + payTypeList[payType].rtnAmt
        tailTxn02RptBody(ws, i, stlm_date, payType, payTypeList[payType].count,
                         payTypeList[payType].transAmt, payTypeList[payType].prodAmt,
                         payTypeList[payType].costAmt, payTypeList[payType].rtnAmt)
        i = i + 1


    tailTxn02RptTail(ws, i, allCount, allTransAmt, allProdAmt, allCostAmt, allRtnAmt)
    i = i + 1

    cursor.close()



def tailTxn02RptBody(ws,i,stlmDate,payType,count,transAmt,prodAmt,costAmt,rtnAmt):
    ws.cell(row=i, column=1).value = stlmDate
    if payType == '00':
        ws.cell(row=i, column=2).value = 'S0����'
    elif payType == '01':
        ws.cell(row=i, column=2).value = 'T1����'
    elif payType == '03':
        ws.cell(row=i, column=2).value = '�����̻���'
    else:
        ws.cell(row=i, column=2).value = '����'
    ws.cell(row=i, column=3).value = count
    ws.cell(row=i, column=4).value = toNumberFmt(fabs(transAmt))
    ws.cell(row=i, column=5).value = toNumberFmt(prodAmt)
    ws.cell(row=i, column=6).value = toNumberFmt(costAmt)
    ws.cell(row=i, column=7).value = toNumberFmt(fabs(transAmt - prodAmt - costAmt))
    ws.cell(row=i, column=8).value = toNumberFmt(fabs(rtnAmt))

def tailTxn02RptTail(ws,i,count,transAmt,prodAmt,costAmt, rtnAmt):
    ws.cell(row=i, column=2).value = '�ܼ�'
    ws.cell(row=i, column=3).value = count
    ws.cell(row=i, column=4).value = toNumberFmt(fabs(transAmt))
    ws.cell(row=i, column=5).value = toNumberFmt(prodAmt)
    ws.cell(row=i, column=6).value = toNumberFmt(costAmt)
    ws.cell(row=i, column=7).value = toNumberFmt(fabs(transAmt - prodAmt - costAmt))
    ws.cell(row=i, column=8).value = toNumberFmt(fabs(rtnAmt))

#����
def tailTxn03RptBody(ws,i,stlm_date,chnlId, type,count,transAmt, rtnAmt):
    typeName = ''
    ws.cell(row=i, column=1).value = stlm_date
    ws.cell(row=i, column=2).value = "%s-%s" % (chnlId, getChnlName(chnlId))
    if type == "00":
        typeName = "S0"
    elif type == "01" :
        typeName = "T1"
    elif type == "03" :
        typeName = "�����̷��󻮿�"
    else:
        typeName = "����"
    ws.cell(row=i, column=3).value = typeName
    ws.cell(row=i, column=4).value = count
    ws.cell(row=i, column=5).value = toNumberFmt(fabs(transAmt))
    if type == "03" :
        ws.cell(row=i, column=6).value = '�����̷���'
    else:
        ws.cell(row=i, column=6).value = '�̻������'
    ws.cell(row=i, column=7).value = toNumberFmt(fabs(rtnAmt))

def handleTxn03RptBody(db, ws, stlm_date):
    # ����ͨ�����Ҷ��˳ɹ��Ĵ���
    global i
    sql = "select trim(DEST_CHNL_ID), pay_type, txn_num, count(*), sum(REAL_TRANS_AMT) " \
          " from TBL_STLM_TXN_BILL_DTL " \
          "where host_date = '%s' and txn_num in ('1801','9164') and check_sta ='1' and " \
          "chnl_id ='A002' group by trim(DEST_CHNL_ID), pay_type, txn_num" % (stlm_date)
    cursor = db.cursor()
    cursor.execute(sql)
    payChnlList = {}
    for ltTxn in cursor:
        key = "%s_%s" % (ltTxn[0],ltTxn[1])
        data = getPayCountData(payChnlList, key)
        if ltTxn[2] == '1801':
            data.transAmt = ltTxn[4]
            data.count = ltTxn[3]
        elif ltTxn[2] == '9164':
            data.rtnAmt = ltTxn[4]
        payChnlList[key] = data


    for key in payChnlList:
        keyData = key.split("_")
        tailTxn03RptBody(ws, i, stlm_date, keyData[0], keyData[1], payChnlList[key].count, payChnlList[key].transAmt, payChnlList[key].rtnAmt)
        i = i + 1

    cursor.close()

def handleTxn03Rpt(db, ws, stlm_date):
    global i
    i = i + 2
    ws.cell(row=i, column=1).value = '����'
    # ����ͷ
    i = i + 1
    ws.cell(row=i, column=1).value = '��������'
    ws.cell(row=i, column=2).value = 'ͨ��'
    ws.cell(row=i, column=3).value = 'S0T1���'
    ws.cell(row=i, column=4).value = '�������'
    ws.cell(row=i, column=5).value = '������'
    ws.cell(row=i, column=6).value = '��������'
    ws.cell(row=i, column=7).value = '�˵����'
    i = i + 1

    handleTxn03RptBody(db, ws, stlm_date)


#POS���̿����ϸ
def handleTxn04Rpt(db, ws, stlm_date):
    handleTxn04RptHead(ws)
    handleTxn04RptBody(db, ws, stlm_date)

def handleTxn04RptHead(ws):
    global i
    i = i + 2
    ws.cell(row=i, column=1).value = 'POS���̿����ϸ'
    #����ͷ
    i = i + 1
    ws.cell(row=i, column=2).value = '��������'
    ws.cell(row=i, column=3).value = '����ʱ��'
    ws.cell(row=i, column=4).value = '���'
    ws.cell(row=i, column=5).value = '�����̻���'
    ws.cell(row=i, column=6).value = '�����ն˺�'
    ws.cell(row=i, column=7).value = '�̻���'
    ws.cell(row=i, column=8).value = '�ն˺�'
    ws.cell(row=i, column=9).value = '�ʺ�'
    ws.cell(row=i, column=10).value = '���׽��'
    ws.cell(row=i, column=11).value = '���������'
    ws.cell(row=i, column=12).value = '���������'
    ws.cell(row=i, column=13).value = 'Ʒ�Ʒ����'
    ws.cell(row=i, column=14).value = '�ʽ����㾻��'

    i = i + 1


def handleTxn04RptBody(db, ws, stlm_date):
    #���ҳ��̿��
    global i
    sql = "select trans_date, trans_time, CHK_STA, key_rsp, CHNL_RETRIVL_REF, GROUP_ID, " \
          " pan " \
          " from TBL_ERR_CHK_TXN_DTL " \
          "where host_date = '%s' and txn_num ='1011' and CHK_STA ='1' order by  CHK_STA" % (stlm_date)
    cursor = db.cursor()
    cursor.execute(sql)
    for ltTxn in cursor:
        transDate = ltTxn[0]
        transTime = ltTxn[1]
        chkSta = ltTxn[2]
        pan = ltTxn[6]
        if chkSta == '1':
            chlTxnKey = ltTxn[4]
            chnlId = ltTxn[5]
            #���������Ϣ
            txnMchtCd, txnTermId, amt, issAmt, swtAmt, prodAmt = getChnlTxn(db, chnlId, chlTxnKey)
            tailTxn04RptBody(ws, i, transDate, transTime, '����', txnMchtCd, txnTermId,
                             '','',pan, amt, issAmt, swtAmt, prodAmt, amt - (issAmt+swtAmt+prodAmt))
        else :
            #��̨�ɹ�
            keyRsp = ltTxn[3]
            mchtCd, termId, amt = getOwnTxn(db, keyRsp)
            tailTxn04RptBody(ws, i, transDate, transTime, '�̿�', '', '',
                             mchtCd, termId, pan, amt, 0.0, 0.0, 0.0, 0.0)
        i = i + 1

    cursor.close()

def tailTxn04RptBody(ws, i, transDate, transTime, type, txnMchtCd, txnTermId,
                     mchtCd, termId, pan, amt, issAmt, swtAmt, prodAmt, stlmAmt) :
    ws.cell(row=i, column=2).value = transDate
    ws.cell(row=i, column=3).value = transTime
    ws.cell(row=i, column=4).value = type
    ws.cell(row=i, column=5).value = txnMchtCd
    ws.cell(row=i, column=6).value = txnTermId
    ws.cell(row=i, column=7).value = mchtCd
    ws.cell(row=i, column=8).value = termId
    ws.cell(row=i, column=9).value = pan
    ws.cell(row=i, column=10).value = amt
    ws.cell(row=i, column=11).value = issAmt
    ws.cell(row=i, column=12).value = swtAmt
    ws.cell(row=i, column=13).value = prodAmt
    ws.cell(row=i, column=14).value = stlmAmt

#�������̿����ϸ
def handleTxn05RptHead(ws):
    global i
    i = i + 2
    ws.cell(row=i, column=1).value = '�������̿����ϸ'
    i = i + 1
    #����ͷ
    ws.cell(row=i, column=2).value = '��������'
    ws.cell(row=i, column=3).value = '���'
    ws.cell(row=i, column=4).value = '�����̻���'
    ws.cell(row=i, column=5).value = '�̻���'
    ws.cell(row=i, column=6).value = '�ն˺�'
    ws.cell(row=i, column=7).value = '�˻���Ϣ'
    ws.cell(row=i, column=8).value = '�������'
    ws.cell(row=i, column=9).value = '�ʽ����㾻��'
    i = i + 1

def tailTxn05RptBody(ws, i, transDate, type, txnMchtCd,
                     mchtCd, termId, pan, amt, stlmAmt) :
    ws.cell(row=i, column=2).value = transDate
    ws.cell(row=i, column=3).value = type
    ws.cell(row=i, column=4).value = txnMchtCd
    ws.cell(row=i, column=5).value = mchtCd
    ws.cell(row=i, column=6).value = termId
    ws.cell(row=i, column=7).value = pan
    ws.cell(row=i, column=8).value = amt
    ws.cell(row=i, column=9).value = stlmAmt


def handleTxn05RptBody(db, ws, stlm_date):
    #���ҳ��̿��
    global i
    sql = "select trans_date, CHK_STA, key_rsp, CHNL_RETRIVL_REF, GROUP_ID, " \
          " pan " \
          " from TBL_ERR_CHK_TXN_DTL " \
          "where host_date = '%s' and txn_num ='1801' and GROUP_ID ='A002' and CHK_STA in ('1','2') order by  CHK_STA" % (stlm_date)
    cursor = db.cursor()
    cursor.execute(sql)
    for ltTxn in cursor:
        transDate = ltTxn[0]
        chkSta = ltTxn[1]
        pan = ltTxn[5]
        if chkSta == '1':
            chlTxnKey = ltTxn[3]
            chnlId = ltTxn[4]
            #���������Ϣ
            txnMchtCd, txnTermId, amt, issAmt, swtAmt, prodAmt = getChnlTxn(db, chnlId, chlTxnKey)
            tailTxn05RptBody(ws, i, transDate, '�̿�', txnMchtCd,
                             '','',pan, amt, amt - issAmt - prodAmt)
        else :
            #��̨�ɹ�
            keyRsp = ltTxn[2]
            mchtCd, termId, amt = getOwnTxn(db, keyRsp)
            tailTxn05RptBody(ws, i, transDate, '����', '',
                             mchtCd, termId, pan, amt, 0.0)
        i = i + 1

    cursor.close()

def handleTxn05Rpt(db, ws, stlm_date):
    handleTxn05RptHead(ws)
    handleTxn05RptBody(db, ws, stlm_date)


def main():
    #���ݿ���������
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'],os.environ['DBPWD'],os.environ['TNSNAME']), encoding='gb18030')
    #��ȡ������
    if len(sys.argv) == 1:
        cursor = db.cursor()
        sql = "select BF_STLM_DATE from TBL_BAT_CUT_CTL"
        cursor.execute(sql)
        x = cursor.fetchone()
        stlm_date = x[0]
        cursor.close()
    else:
        stlm_date = sys.argv[1]

    print('hostDate %s rpt begin' % stlm_date)

    #����ָ���ļ�
    wb = Workbook()
    ws = wb.active
    handleTxn01Rpt(db, ws, stlm_date)
    handleTxn02Rpt(db, ws, stlm_date)
    handleTxn03Rpt(db, ws, stlm_date)
    handleTxn04Rpt(db, ws, stlm_date)
    handleTxn05Rpt(db, ws, stlm_date)

    filePath = '%s/%s/' % (os.environ['RPT7HOME'],stlm_date)
    filename = filePath + 'AcqStlmCheckFile02_%s.xlsx' % stlm_date
    wb.save(filename)

if __name__ == '__main__':
    main()

