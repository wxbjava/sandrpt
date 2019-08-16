#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#间联系统余额报表,按机构每个机构一个

import cx_Oracle
import sys
import os
from math import fabs
from openpyxl.workbook import Workbook
from utl.common import *

class MchtBalance:
    def __init__(self, insIdCd):
        self.initAmt = 0.0
        self.finalAmt = 0.0
        self.insIdCd = insIdCd
        self.txnCount = 0        #商户交易笔数
        self.txnAmt = 0.0        #商户交易总金额
        self.txnCost = 0.0       #总成本
        self.errAmt = 0.0        #差错
        self.mchtFee = 0.0       #商户手续费
        self.mchtStlmAmt = 0.0   #商户结算费
        self.payTxnCount = 0     #代付笔数
        self.payTxnAmt = 0.0     #代付金额
        self.payUnknownCount = 0.0   #未知代付笔数
        self.payUnknownAmt = 0.0  #未知代付金额
        self.payPayTxnRtn = 0.0   #代付退回金额

    def __get_balance_amt(self, db, stlmDate):
        sql = "select sum(MCHT_A_PREV_BAL_AT + MCHT_B_PREV_BAL_AT - MCHT_C_PREV_BAL_AT) " \
              "from TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (getLastDay(stlmDate), self.insIdCd)
        cursor = db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.initAmt = toNumberFmt(x[0]/100)
        sql = "select sum(MCHT_A_PREV_BAL_AT + MCHT_B_PREV_BAL_AT - MCHT_C_PREV_BAL_AT) " \
              "from TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (stlmDate, self.insIdCd)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.finalAmt = toNumberFmt(x[0]/100)
        cursor.close()

    def __get_succ_txn(self, db, stlmDate):
        sql = "select count(*), txn_num, nvl(sum(real_trans_amt),0), nvl(sum(ISS_FEE+SWT_FEE+PROD_FEE),0), " \
              "nvl(sum(ERR_FEE),0), nvl(sum(mcht_fee),0) " \
              "from TBL_STLM_TXN_BILL_DTL where " \
              "CHECK_STA ='1' and host_date = '%s' " \
              "and ins_id_cd = '%s' group by txn_num" % (stlmDate, self.insIdCd)
        print(sql)
        cursor = db.cursor()
        cursor.execute(sql)
        for ltTxn in cursor:
            if ltTxn[1] == '1011':
                # 消费
                self.txnCount = ltTxn[0]
                self.txnAmt = toNumberFmt(ltTxn[2])
                self.txnCost = toNumberFmt(ltTxn[3])
                self.mchtFee = toNumberFmt(ltTxn[5])
            elif ltTxn[1] == '1801':
                # 代付
                self.payTxnCount = ltTxn[0]
                self.payTxnAmt = fabs(ltTxn[2])

        #计算代理商分润
        sql = "select count(*), nvl(sum(trans_amt/100),0) from tbl_acq_txn_log where host_date ='%s' and " \
              "txn_num ='1801' and substrb(ADDTNL_DATA,1,2) in ('04','05','06') and " \
              "trans_state ='1' and company_cd ='%s'" % (stlmDate, self.insIdCd)
        cursor.execute(sql)
        x = cursor.fetchone()
        self.payTxnCount = self.payTxnCount - x[0]
        self.payTxnAmt = toNumberFmt(self.payTxnAmt - x[1])

        cursor.close()
        self.mchtStlmAmt = toNumberFmt(self.txnAmt - self.errAmt - self.mchtFee)

    def __get_oth_txn(self, db, dbacc, stlmDate):
        sql = "select count(*), sum(a.txn_amt) from tbl_err_chk_txn_dtl a " \
              "left join tbl_mcht_inf b on a.CARD_ACCP_ID = b.mcht_cd " \
              "left join tbl_acq_txn_log c on a.key_rsp = c.key_rsp " \
              "where a.host_date ='%s' and a.chk_sta='4' and b.company_cd = '%s' " \
              " and a.txn_num ='1801' and substr(c.ADDTNL_DATA,1,2) = '02'" % (stlmDate, self.insIdCd)
        print(sql)
        cursor = db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.payUnknownCount = toNumberFmt(x[0])
            self.payUnknownAmt = toNumberFmt(x[1])
        cursor.close()

        #查找非当日代付退回记录
        sql = "select sum(a.TXN_AT - a.TXN_FEE_AT)/100 from " \
              "(select * from t_txn_log where host_date ='%s' and TXN_NUM ='801012') a " \
              "left join (select * from t_txn_log where TXN_NUM='801011') b " \
              "on a.txn_key = b.txn_key where a.host_date != b.host_date and a.ACCP_BRH_ID = '%s' and " \
              "length(trim(a.ext_acct_id)) = 15" % (stlmDate, self.insIdCd)
        print(sql)
        cursor = dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.payPayTxnRtn = toNumberFmt(x[0])
        cursor.close()



    def getAcctInfo(self, db, dbacc, stlmDate):
        self.__get_balance_amt(db, stlmDate)
        self.__get_succ_txn(db, stlmDate)
        self.__get_oth_txn(db, dbacc, stlmDate)

class AgentBalance:
    def __init__(self, insIdCd, dbbat, dbacc, stlmDate):
        self.insIdCd = insIdCd
        self.dbbat = dbbat
        self.dbacc = dbacc
        self.stlmDate = stlmDate
        self.agentInitAmt = 0.0
        self.agentFinalAmt = 0.0
        self.agentPay = 0.0
        self.agentPayUnknownCount = 0
        self.agentPayUnknownAmt = 0
        self.agentPayUnknownRtn = 0.0
        self.agentIncome = 0.0
        self.agentDelayIncome = 0.0
        self.companyInitAmt = 0.0
        self.companyFinalAmt = 0.0
        self.companyIncome = 0.0
        self.companyPay = 0.0
        self.companyDelayIncome = 0.0
        self.agentAcctId = self.__get_agent_acct_id()
        self.companyAcctId = self.__get_company_acct_id()


    def __get_agent_acct_id(self):
        sql = "select ACCT_ID from t_acct_map where ext_acct_id ='%sA' and EXT_ACCT_TYPE ='0000000B'" % self.insIdCd
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        cursor.close()
        if x is not None:
            return x[0]
        else:
            return '0'

    def __get_company_acct_id(self):
        sql = "select ACCT_ID from t_acct_map where ext_acct_id ='%sB' and EXT_ACCT_TYPE ='0000000B'" % self.insIdCd
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        cursor.close()
        if x is not None:
            return x[0]
        else:
            return '0'


    def __get_agent_balance_amt(self):
        sql = "select sum(INS_B_PREV_BAL_AT - INS_C_PREV_BAL_AT) " \
              "from TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (getLastDay(self.stlmDate), self.insIdCd)
        cursor = self.dbbat.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.agentInitAmt = toNumberFmt(x[0]/100)

        sql = "select sum(INS_B_PREV_BAL_AT - INS_C_PREV_BAL_AT) " \
              "from TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (self.stlmDate, self.insIdCd)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.agentFinalAmt = toNumberFmt(x[0]/100)

        sql = "select sum(INS_B_PREV_BAL_AT - INS_B_PREV_AVAIL_AT) from " \
              "TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (self.stlmDate, self.insIdCd)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.agentFinalLockAmt = toNumberFmt(x[0] / 100)
        else :
            self.agentFinalLockAmt = 0

        cursor.close()

    def __get_company_balance_amt(self):
        sql = "select sum(ACQ_PREV_BAL_AT) " \
              "from TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (getLastDay(self.stlmDate), self.insIdCd)
        cursor = self.dbbat.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.companyInitAmt = toNumberFmt(x[0]/100)

        sql = "select sum(ACQ_PREV_BAL_AT) " \
              "from TBL_SAND_BALANCE_INF where host_date ='%s' and INS_ID_CD ='%s'" % (self.stlmDate, self.insIdCd)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.companyFinalAmt = toNumberFmt(x[0]/100)
        cursor.close()

    def __get_agent_income(self):
        sql = "select sum(ALL_PROFITS) from " \
              "TBL_INS_PROFITS_TXN_SUM where " \
              "host_date <'%s' and INS_ID_CD ='%s' and to_char(REC_UPD_TS, 'YYYYMMDD') = '%s'" % \
              (self.stlmDate, self.insIdCd, self.stlmDate)
        print(sql)
        cursor = self.dbbat.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.agentIncome = toNumberFmt(x[0])
        cursor.close()

        #计算交易日对应收入
        sql = "select sum(ALL_PROFITS) from tbl_ins_profits_txn_sum where " \
              "host_date <= '%s' and INS_ID_CD ='%s' and CHARGE_STA != '2'" % (self.stlmDate, self.insIdCd)
        cursor = self.dbbat.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.agentDelayIncome = toNumberFmt(x[0])
        cursor.close()

    def __get_agent_pay(self):
        sql = "select sum(TXN_AT/100) from t_txn_dtl " \
              "where ACCEPT_DT ='%s' and acct_id ='%s' " \
              "and ACCT_TYPE ='00000002' and INT_TXN_CD in ('01005','01033')" % \
              (self.stlmDate, self.agentAcctId)
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.agentPay = toNumberFmt(x[0])
        sql = "select sum(TXN_AT/100) from t_txn_dtl " \
              "where ACCEPT_DT ='%s' and acct_id ='%s' " \
              "and ACCT_TYPE ='00000002' and INT_TXN_CD in ('01010','01034') and txn_part_cd like '%s%%'" % \
              (self.stlmDate, self.agentAcctId, self.stlmDate[4:8])
        cursor.execute(sql)
        x = cursor.fetchone()
        cursor.close()
        if x is not None:
            self.agentPay = self.agentPay - toNumberFmt(x[0])

    def __get_agent_pay_unknown(self):
        sql = "select count(*), sum(a.txn_amt) from tbl_err_chk_txn_dtl a " \
              "left join tbl_mcht_inf b on a.CARD_ACCP_ID = b.mcht_cd " \
              "left join tbl_acq_txn_log c on a.key_rsp = c.key_rsp " \
              "where a.host_date ='%s' and a.chk_sta='4' and c.company_cd = '%s' " \
              " and a.txn_num ='1801' and substr(c.ADDTNL_DATA,1,2) in ('04','05','06')" % (self.stlmDate, self.insIdCd)
        print(sql)
        cursor = self.dbbat.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.agentPayUnknownCount = toNumberFmt(x[0])
            self.agentPayUnknownAmt = toNumberFmt(x[1])
        cursor.close()

        # 查找非当日代付退回记录
        sql = "select sum(a.TXN_AT - a.TXN_FEE_AT)/100 from " \
              "(select * from t_txn_log where host_date ='%s' and TXN_NUM in ('801010','801034')) a " \
              "left join (select * from t_txn_log where TXN_NUM in ('801005','801033')) b " \
              "on a.txn_key = b.txn_key where a.host_date != b.host_date and a.ACCP_BRH_ID = '%s' and " \
              "a.acct_id = '%s'" % (self.stlmDate, self.insIdCd, self.agentAcctId)
        print(sql)
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.agentPayUnknownRtn = toNumberFmt(x[0])
        cursor.close()

    def __get_agent_lockamt(self):
        #冻结
        sql = "select sum(LOCK_AT)/100 from T_TXN_LOCK where " \
              "host_date ='%s' and TXN_TYPE ='01' and acct_id ='%s'" % (self.stlmDate, self.agentAcctId)
        self.agentLockAmt = 0
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.agentLockAmt = toNumberFmt(x[0])

        #解冻
        sql = "select sum(LOCK_AT)/100 from T_TXN_LOCK where " \
              "host_date ='%s' and TXN_TYPE ='02' and acct_id ='%s'" % (self.stlmDate, self.agentAcctId)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.agentLockAmt = toNumberFmt(self.agentLockAmt - x[0])

        cursor.close()

    def __get_company_income(self):
        sql = "select sum(TXN_AT/100) from t_txn_dtl " \
              "where ACCEPT_DT ='%s' and acct_id ='%s' " \
              "and ACCT_TYPE ='00000002' and INT_TXN_CD='01004'" % \
              (self.stlmDate, self.companyAcctId)
        print(sql)
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.companyIncome = toNumberFmt(x[0])
        sql = "select nvl(sum(TXN_AT/100),0) from t_txn_dtl " \
              "where ACCEPT_DT ='%s' and acct_id ='%s' " \
              "and ACCT_TYPE ='00000002' and INT_TXN_CD='01003' and txn_part_cd not like '%%核销%%' " % \
              (self.stlmDate, self.companyAcctId)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.companyIncome = toNumberFmt(self.companyIncome - x[0])
        cursor.close()

        # 计算交易日对应收入
        sql = "select sum(ALL_PROFITS) from TBL_SAND_ACQ_PROFITS where " \
                "host_date = '%s' and INS_ID_CD ='%s'" % (self.stlmDate, self.insIdCd)
        cursor = self.dbbat.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.companyDelayIncome = toNumberFmt(x[0])
        cursor.close()

    def __get_company_pay(self):
        sql = "select sum(TXN_AT/100) from t_txn_dtl " \
              "where ACCEPT_DT ='%s' and acct_id ='%s' " \
              "and ACCT_TYPE ='00000002' and CR_DB_CD='0' and txn_part_cd like '%%核销%%'" % \
              (self.stlmDate, self.companyAcctId)
        cursor = self.dbacc.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        cursor.close()
        if x is not None:
            self.companyPay = toNumberFmt(x[0])



    def getAcctInfo(self):
        self.__get_agent_balance_amt()
        self.__get_agent_income()
        self.__get_agent_pay()
        self.__get_agent_lockamt()
        self.__get_company_balance_amt()
        self.__get_company_income()
        self.__get_company_pay()
        self.__get_agent_pay_unknown()

def insertDb(stlmDate, db, mchtBal, agentBal):
    sql = "insert into TBL_RPT_INS_BALANCE_INF (host_date,ins_id_cd,mcht_init_at,txn_count,txn_amt,txn_cost,err_amt," \
          "mcht_fee,mcht_stlm_amt,pay_txn_count,pay_txn_amt,pay_unknown_count,pay_unknown_amt,pay_txn_rtn_amt," \
          "mcht_sys_diff_at,mcht_final_at,agent_init_at,agent_income,agent_err_amt,agent_pay,agent_pay_unknown_count," \
          "agent_pay_unknown_amt,agent_pay_txn_rtn_amt,agent_delay_income,agent_lock_amt,agent_final_lock_amt," \
          "agent_final_at,company_init_at,company_income,company_pay,company_delay_income,company_final_at) values (" \
          ":1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15,:16,:17,:18,:19," \
          ":20,:21,:22,:23,:24,:25,:26,:27,:28,:29,:30,:31,:32)"
    cursor = db.cursor()
    cursor.prepare(sql)
    param = (stlmDate, mchtBal.insIdCd, mchtBal.initAmt, mchtBal.txnCount,mchtBal.txnAmt, mchtBal.txnCost,
             mchtBal.errAmt, mchtBal.mchtFee,mchtBal.mchtStlmAmt, mchtBal.payTxnCount, mchtBal.payTxnAmt,
                                            mchtBal.payUnknownCount, mchtBal.payUnknownAmt, mchtBal.payPayTxnRtn, 0, mchtBal.finalAmt,
                                            agentBal.agentInitAmt, agentBal.agentIncome, 0, agentBal.agentPay,
                                            agentBal.agentPayUnknownCount, agentBal.agentPayUnknownAmt, agentBal.agentPayUnknownRtn,
                                           agentBal.agentDelayIncome, agentBal.agentLockAmt, agentBal.agentFinalLockAmt,
                                           agentBal.agentFinalAmt,
                                            agentBal.companyInitAmt, agentBal.companyIncome, agentBal.companyPay,
                                            agentBal.companyDelayIncome, agentBal.companyFinalAmt)
    cursor.execute(None, param)
    cursor.close()


def genRptFunc(stlmDate, db, ws, mchtBal, agentBal):
    i = 1
    ws.cell(row=i, column=8).value = '间联系统余额报表'
    i = i + 1
    # 报表头
    ws.cell(row=i, column=1).value = '交易日期'
    ws.cell(row=i, column=2).value = '商户期初余额'
    ws.cell(row=i, column=3).value = '交易笔数'
    ws.cell(row=i, column=4).value = '交易金额'
    ws.cell(row=i, column=5).value = '总成本'
    ws.cell(row=i, column=6).value = '差错费'
    ws.cell(row=i, column=7).value = '手续费'
    ws.cell(row=i, column=8).value = '商户应出账'
    ws.cell(row=i, column=9).value = '代付笔数'
    ws.cell(row=i, column=10).value = '代付金额'
    ws.cell(row=i, column=11).value = '代付未知笔数'
    ws.cell(row=i, column=12).value = '代付未知金额'
    ws.cell(row=i, column=13).value = '未知代付退回金额'
    ws.cell(row=i, column=14).value = '商户期末余额'
    ws.cell(row=i, column=15).value = '机构合作商期初余额'
    ws.cell(row=i, column=16).value = '机构合作商收入'
    ws.cell(row=i, column=17).value = '机构合作商差错费'
    ws.cell(row=i, column=18).value = '机构合作商划款'
    ws.cell(row=i, column=19).value = '机构合作商划款未知笔数'
    ws.cell(row=i, column=20).value = '机构合作商划款未知金额'
    ws.cell(row=i, column=21).value = '机构合作商划款未知金额退回'
    ws.cell(row=i, column=22).value = '机构合作商待入账收入'
    ws.cell(row=i, column=23).value = '机构合作商冻结解冻金额'
    ws.cell(row=i, column=24).value = '机构合作商冻结总额'
    ws.cell(row=i, column=25).value = '机构合作商期末余额'
    ws.cell(row=i, column=26).value = '杉德收入期初余额'
    ws.cell(row=i, column=27).value = '杉德收入'
    ws.cell(row=i, column=28).value = '杉德收入划款'
    ws.cell(row=i, column=29).value = '杉德待入账收入'
    ws.cell(row=i, column=30).value = '杉德收入未结余额'

    #值
    i = i + 1
    ws.cell(row=i, column=1).value = stlmDate
    ws.cell(row=i, column=2).value = mchtBal.initAmt
    ws.cell(row=i, column=3).value = mchtBal.txnCount
    ws.cell(row=i, column=4).value = mchtBal.txnAmt
    ws.cell(row=i, column=5).value = mchtBal.txnCost
    ws.cell(row=i, column=6).value = mchtBal.errAmt
    ws.cell(row=i, column=7).value = mchtBal.mchtFee
    ws.cell(row=i, column=8).value = mchtBal.mchtStlmAmt
    ws.cell(row=i, column=9).value = mchtBal.payTxnCount
    ws.cell(row=i, column=10).value = mchtBal.payTxnAmt
    ws.cell(row=i, column=11).value = mchtBal.payUnknownCount
    ws.cell(row=i, column=12).value = mchtBal.payUnknownAmt
    ws.cell(row=i, column=13).value = mchtBal.payPayTxnRtn
    ws.cell(row=i, column=14).value = mchtBal.finalAmt
    ws.cell(row=i, column=15).value = agentBal.agentInitAmt
    ws.cell(row=i, column=16).value = agentBal.agentIncome
    ws.cell(row=i, column=17).value = 0
    ws.cell(row=i, column=18).value = agentBal.agentPay
    ws.cell(row=i, column=19).value = agentBal.agentPayUnknownCount
    ws.cell(row=i, column=20).value = agentBal.agentPayUnknownAmt
    ws.cell(row=i, column=21).value = agentBal.agentPayUnknownRtn
    ws.cell(row=i, column=22).value = agentBal.agentDelayIncome
    ws.cell(row=i, column=23).value = agentBal.agentLockAmt
    ws.cell(row=i, column=24).value = agentBal.agentFinalLockAmt
    ws.cell(row=i, column=25).value = agentBal.agentFinalAmt
    ws.cell(row=i, column=26).value = agentBal.companyInitAmt
    ws.cell(row=i, column=27).value = agentBal.companyIncome
    ws.cell(row=i, column=28).value = agentBal.companyPay
    ws.cell(row=i, column=29).value = agentBal.companyDelayIncome
    ws.cell(row=i, column=30).value = agentBal.companyFinalAmt

    insertDb(stlmDate, db, mchtBal, agentBal)

def main():
    # 数据库连接配置
    dbbat = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),encoding='gb18030')
    dbacc = cx_Oracle.connect('%s/%s@%s' % (os.environ['ACCDBUSER'], os.environ['ACCDBPWD'], os.environ['TNSNAME']),
                              encoding='gb18030')
    # 获取清算日
    if len(sys.argv) == 1:
        cursor = dbbat.cursor()
        sql = "select BF_STLM_DATE from TBL_BAT_CUT_CTL"
        cursor.execute(sql)
        x = cursor.fetchone()
        stlm_date = x[0]
        cursor.close()
    else:
        stlm_date = sys.argv[1]

    print('hostDate %s genRptAcqBalance begin' % stlm_date)

    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_date)

    #查找机构
    sql = "select trim(INS_ID_CD) from TBL_INS_INF where INS_TP ='01'"
    cursor = dbbat.cursor()
    cursor.execute(sql)
    for ltData in cursor:
        if ltData[0] is not None:
            #查看信息
            insIdCd = ltData[0]
            mchtBal = MchtBalance(insIdCd)
            mchtBal.getAcctInfo(dbbat, dbacc, stlm_date)
            agentBal = AgentBalance(insIdCd, dbbat, dbacc, stlm_date)
            agentBal.getAcctInfo()
            filename = filePath + 'AcqBalanceInf_%s_%s.xlsx' % (insIdCd,stlm_date)
            wb = Workbook()
            ws = wb.active
            genRptFunc(stlm_date, dbbat, ws, mchtBal, agentBal)
            wb.save(filename)
            wb.close()
    cursor.close()
    dbbat.commit()



if __name__ == '__main__':
    main()
