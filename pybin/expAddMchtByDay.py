#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#导出每日新增商户

import cx_Oracle
import sys
import os
from openpyxl.workbook import Workbook


i = 1

def newAddMchtFileHead(ws):
    global i
    ws.cell(row=i, column=1).value = '代理商号'
    ws.cell(row=i, column=2).value = '代理商名'
    ws.cell(row=i, column=3).value = '商户号'
    ws.cell(row=i, column=4).value = '商户名'
    ws.cell(row=i, column=5).value = '注册日期'
    ws.cell(row=i, column=6).value = '结算账户名'
    ws.cell(row=i, column=7).value = '结算帐号'
    ws.cell(row=i, column=8).value = '身份证'

    i = i + 1

def newAddMchtFileBody(ws, company_cd, company_nm, mcht_cd, mcht_cn, crt_dt, acct_nm, acct_no, mcht_certif_no):
    global i
    ws.cell(row=i, column=1).value = company_cd
    ws.cell(row=i, column=2).value = company_nm
    ws.cell(row=i, column=3).value = mcht_cd
    ws.cell(row=i, column=4).value = mcht_cn
    ws.cell(row=i, column=5).value = crt_dt
    ws.cell(row=i, column=6).value = acct_nm
    ws.cell(row=i, column=7).value = acct_no
    ws.cell(row=i, column=8).value = mcht_certif_no
    i = i + 1

def getInsName(db, ins_id_cd):
    sql = "select INS_NAME from tbl_ins_inf where ins_id_cd = '%s'" % ins_id_cd
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    if x is not None:
        result = x[0]
    else:
        result = '未知代理商'
    cursor.close()
    return result

def getMchtAcct(db, mcht_cd):
    sql = "select MCHT_STLM_C_NM,MCHT_STLM_C_ACCT from tbl_mcht_acct_inf where mcht_cd = '%s'" % mcht_cd
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    if x is not None:
        return x[0], x[1]
    else:
        return ' ',' '

def main():
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),encoding='gb18030')
    stlm_date = sys.argv[1]
    print('hostDate %s expAddMchtByDay begin' % stlm_date)
    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_date)
    filename = filePath + 'AddMchtByDay_%s.xlsx' % (stlm_date)

    sql = "select mcht_cd, company_cd, MCHT_CERTIF_NO, to_char(REC_CRT_TS, 'YYYYMMDD'), trim(mcht_cn) from " \
          "tbl_mcht_inf where to_char(REC_CRT_TS, 'YYYYMMDD') = '%s'" % stlm_date
    print(sql)
    wb = Workbook()
    ws = wb.active
    newAddMchtFileHead(ws)
    cursor = db.cursor()
    cursor.execute(sql)
    for ltData in cursor:
        company_cd = ltData[1]
        company_nm = getInsName(db, company_cd)
        mcht_cd = ltData[0]
        mcht_certif_no = ltData[2]
        crt_dt = ltData[3]
        mcht_cn = ltData[4]
        acct_nm, acct_no = getMchtAcct(db, mcht_cd)
        newAddMchtFileBody(ws, company_cd, company_nm, mcht_cd, mcht_cn, crt_dt, acct_nm, acct_no, mcht_certif_no)


    cursor.close()
    wb.save(filename)
    wb.close()
    print('hostDate %s expAddMchtByDay end' % stlm_date)


if __name__ == '__main__':
    main()