#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#�����ɱ�����

import cx_Oracle
import sys
import os
from math import fabs
from openpyxl.workbook import Workbook
from utl.common import *
from utl.gldict import *

def getMchtName(mchtCd, db):
    sql = "select MCHT_NAME from TBL_OBJ_MCHT_INF where mcht_cd ='%s'" % mchtCd
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    if x is None:
        return "δ֪�̻���"
    else:
        return x[0]

def getStlmMd(stlmMd):
    if stlmMd == '0':
        return "S0"
    elif stlmMd == '1':
        return "T1"
    else:
        return "δ֪"

def getMchtType(mchtType):
    if mchtType == '1':
        return "��׼"
    elif mchtType == '2':
        return "�Ż�"
    else:
        return "����"

def getCardType(cardType):
    if cardType == '00':
        return "����"
    elif cardType == '01':
        return "���"
    else:
        return "δ֪"


def newSandCostFileHead(ws):
    i = 1
    ws.cell(row=i, column=1).value = '��������'
    ws.cell(row=i, column=2).value = '�̻���'
    ws.cell(row=i, column=3).value = '�̻���'
    ws.cell(row=i, column=4).value = '������'
    ws.cell(row=i, column=5).value = '��Ŀ��ʶ'
    ws.cell(row=i, column=6).value = '�̻�����'
    ws.cell(row=i, column=7).value = '��/��'
    ws.cell(row=i, column=8).value = '���ױ���'
    ws.cell(row=i, column=9).value = '���׽��'
    ws.cell(row=i, column=10).value = '�̻�������'
    ws.cell(row=i, column=11).value = '�����з����'
    ws.cell(row=i, column=12).value = '��������ת�ӷ�'
    ws.cell(row=i, column=13).value = 'Ʒ�Ʒ����'
    ws.cell(row=i, column=14).value = '�ܳɱ�'
    ws.cell(row=i, column=15).value = '��������'
    ws.cell(row=i, column=16).value = '���˽��'
    ws.cell(row=i, column=17).value = '�̻������ʽ�'
    ws.cell(row=i, column=18).value = '�յ�����'
    ws.cell(row=i, column=19).value = '�ܲ�����'
    ws.cell(row=i, column=20).value = '�ֹ�˾����'
    ws.cell(row=i, column=21).value = '��������'

def tailSandCostBody(ws,i,ltData, db):
    ws.cell(row=i, column=1).value = ltData[1]
    ws.cell(row=i, column=2).value = ltData[2]
    ws.cell(row=i, column=3).value = getMchtName(ltData[2], db)
    ws.cell(row=i, column=4).value = getStlmMd(ltData[3])
    ws.cell(row=i, column=5).value = getItemName(ltData[4])
    ws.cell(row=i, column=6).value = getMchtType(ltData[5])
    ws.cell(row=i, column=7).value = getCardType(ltData[6])
    ws.cell(row=i, column=8).value = ltData[7]
    ws.cell(row=i, column=9).value = toNumberFmt(ltData[8])
    ws.cell(row=i, column=10).value = toNumberFmt(ltData[9])
    ws.cell(row=i, column=11).value = toNumberFmt(ltData[10])
    ws.cell(row=i, column=12).value = toNumberFmt(ltData[11])
    ws.cell(row=i, column=13).value = toNumberFmt(ltData[12])
    ws.cell(row=i, column=14).value = toNumberFmt(ltData[13])
    ws.cell(row=i, column=15).value = toNumberFmt(ltData[14])
    ws.cell(row=i, column=16).value = toNumberFmt(ltData[15])
    ws.cell(row=i, column=17).value = toNumberFmt(ltData[16])
    ws.cell(row=i, column=18).value = toNumberFmt(ltData[17])
    ws.cell(row=i, column=19).value = toNumberFmt(ltData[18])
    ws.cell(row=i, column=20).value = toNumberFmt(ltData[19])
    ws.cell(row=i, column=21).value = toNumberFmt(ltData[20])

def main():
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),encoding='gb18030')
    if len(sys.argv) == 1:
        cursor = db.cursor()
        sql = "select BF_STLM_DATE from TBL_BAT_CUT_CTL"
        cursor.execute(sql)
        x = cursor.fetchone()
        stlm_date = x[0]
        cursor.close()
    else:
        stlm_date = sys.argv[1]

    print('hostDate %s expRptSandCost begin' % stlm_date)
    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_date)
    if (len(stlm_date) == 8) :
        sql = "select trim(INS_ID_CD),HOST_DATE,MCHT_CD,STLM_MD,ITERM_ID, " \
              "MCHT_TYPE,CARD_TYPE,TXN_SUM,TXN_AMT,MCHT_FEE,ISS_FEE,SWT_FEE,PROD_FEE," \
              "ALL_COST_FEE,ERR_FEE,ACC_IN_AMT,MCHT_STLM_AMT,ACQ_AMT,HEAD_AMT,BRA_AMT,AGENT_AMT " \
              "from TBL_SAND_COST_FILE_DTL where host_date ='%s' order by INS_ID_CD" % stlm_date
    else:
        sql = "select trim(INS_ID_CD),HOST_DATE,MCHT_CD,STLM_MD,ITERM_ID, " \
              "MCHT_TYPE,CARD_TYPE,TXN_SUM,TXN_AMT,MCHT_FEE,ISS_FEE,SWT_FEE,PROD_FEE," \
              "ALL_COST_FEE,ERR_FEE,ACC_IN_AMT,MCHT_STLM_AMT,ACQ_AMT,HEAD_AMT,BRA_AMT,AGENT_AMT " \
              "from TBL_SAND_COST_FILE_DTL where host_date like '%s%%' order by INS_ID_CD, host_date" % stlm_date
    print(sql)
    cursor = db.cursor()
    cursor.execute(sql)
    insIdCdTmp = ''
    for ltData in cursor:
        if insIdCdTmp == '':
            insIdCdTmp = ltData[0]
            filename = filePath + 'Sand_Cost_%s_%s.xlsx' % (ltData[0], stlm_date)
            wb = Workbook()
            ws = wb.active
            newSandCostFileHead(ws)
            i = 2
        if insIdCdTmp != ltData[0]:
            #�ر�ǰһ�������ļ�
            wb.save(filename)
            wb.close()
            #�´����������ļ�
            insIdCdTmp = ltData[0]
            filename = filePath + 'Sand_Cost_%s_%s.xlsx' % (ltData[0], stlm_date)
            wb = Workbook()
            ws = wb.active
            newSandCostFileHead(ws)
            i = 2
        #д���ļ�
        tailSandCostBody(ws,i,ltData, db)
        i = i + 1
    if insIdCdTmp != '':
        wb.save(filename)
        wb.close()
    cursor.close()
    print('hostDate %s expRptSandCost end' % stlm_date)


if __name__ == '__main__':
    main()