#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#�±���-ͨ�������ɱ�

import cx_Oracle
import sys
import os
from math import fabs
from openpyxl.workbook import Workbook
from utl.common import *
from utl.gldict import *

i = 1

def newPayTxnCostFileHead(ws):
    global i
    ws.cell(row=i, column=1).value = '�·�'
    ws.cell(row=i, column=2).value = 'ͨ�����'
    ws.cell(row=i, column=3).value = 'ͨ������'
    ws.cell(row=i, column=4).value = '�����ܽ��'
    ws.cell(row=i, column=5).value = '��������'
    ws.cell(row=i, column=6).value = '�����ɱ�'
    i = i + 1

def newPayTxnCostFileBody(ws, stlmMonth, chnlId, count, allAmt, allCost):
    global i
    ws.cell(row=i, column=1).value = stlmMonth
    ws.cell(row=i, column=2).value = chnlId
    ws.cell(row=i, column=3).value = getChnlName(chnlId)
    ws.cell(row=i, column=4).value = toNumberFmt(allAmt)
    ws.cell(row=i, column=5).value = count
    ws.cell(row=i, column=6).value = toNumberFmt(allCost)
    i = i + 1


def main():
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),encoding='gb18030')
    stlm_month = sys.argv[1]
    print('hostDate %s expRptPayTxnCost begin' % stlm_month)
    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_month)
    filename = filePath + 'Sand_PayTxn_%s.xlsx' % (stlm_month)

    sql = "select TXN_NUM, REAL_TRANS_AMT, DEST_CHNL_ID from TBL_STLM_TXN_BILL_DTL where stlm_date like '%s%%' and  " \
          "CHNL_ID = 'A002' order by DEST_CHNL_ID" % stlm_month
    print(sql)
    wb = Workbook()
    ws = wb.active
    newPayTxnCostFileHead(ws)
    cursor = db.cursor()
    cursor.execute(sql)
    chnlId = ''
    for ltData in cursor:
        if chnlId == '':
            chnlId = ltData[2]
            count = 0
            allamt = 0
            allcost = 0
        elif chnlId != ltData[2]:
            #�Ǽ��ļ�
            newPayTxnCostFileBody(ws, stlm_month, chnlId, count, allamt, allcost)
            chnlId = ltData[2]
            count = 0
            allamt = 0
            allcost = 0

        if fabs(ltData[1]) <= 1000:
            cost = 0.05
        elif fabs(ltData[1]) <= 50000:
            cost = fabs(ltData[1]) * 0.005 / 100
        else:
            cost = 3.5

        if ltData[0] == '1801':
            count = count + 1
            allamt = allamt + fabs(ltData[1])
            allcost = allcost + cost
        else:
            count = count - 1
            allamt = allamt - fabs(ltData[1])
            allcost = allcost - cost

    if chnlId != '':
        newPayTxnCostFileBody(ws, stlm_month, chnlId, count, allamt, allcost)

    cursor.close()
    wb.save(filename)
    wb.close()
    print('hostDate %s expRptPayTxnCost end' % stlm_month)


if __name__ == '__main__':
    main()