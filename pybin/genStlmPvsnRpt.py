#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#�����𱨱�,���ɱ����Լ���¼��ĩ������TBL_STLM_PVSN_RPT

import os
import sys
import cx_Oracle
from openpyxl.workbook import Workbook
from utl.common import *
from utl.gldict import *
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


class rptFile():
    def __init__(self, ws):
        self.ws = ws
        self.iCurr = 1

    def setTitle(self, sheetName):
        self.ws.title = sheetName

    def head(self):
        self.ws.cell(row=1, column=3).value = '�̻��������'
        self.ws.cell(row=self.iCurr, column=3).alignment = Alignment(wrap_text=True)
        self.ws.cell(row=1, column=4).value = '�����쳣����'
        self.ws.cell(row=self.iCurr, column=4).alignment = Alignment(wrap_text=True)
        self.ws.cell(row=1, column=5).value = '���յ���'
        self.ws.cell(row=1, column=6).value = 'δ����˾����'
        self.ws.cell(row=self.iCurr, column=6).alignment = Alignment(wrap_text=True)
        self.ws.cell(row=1, column=7).value = 'δ������������'
        self.ws.cell(row=self.iCurr, column=7).alignment = Alignment(wrap_text=True)
        self.ws.cell(row=1, column=8).value = 'Ӧ������'
        self.ws.cell(row=1, column=9).value = 'Ӧ��������������'
        self.ws.cell(row=self.iCurr, column=9).alignment = Alignment(wrap_text=True)
        self.ws.cell(row=1, column=10).value = '���д��'
        self.ws.cell(row=1, column=11).value = '����δ����'
        self.ws.cell(row=self.iCurr, column=11).alignment = Alignment(wrap_text=True)
        self.ws.cell(row=1, column=12).value = '��ض���'
        self.ws.cell(row=1, column=13).value = '��������'
        self.iCurr = 2

    def recordInitAmt(self, initAcct):
        self.__recordDtlAmt(initAcct, '�ڳ�')

    #����-�������ף����������˵���
    def recordTodayTxn(self, initAcct, mchtStlmAmt, companyIncome,
                       insProfits, chnlAmt):
        self.ws.cell(row=self.iCurr, column=1).value = initAcct.stlmDate
        self.ws.cell(row=self.iCurr, column=2).value = '����-�������ף����������˵���'
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt + mchtStlmAmt)
        self.ws.cell(row=self.iCurr, column=3).value = mchtStlmAmt
        initAcct.companyIncome = toNumberFmt(initAcct.companyIncome + companyIncome)
        self.ws.cell(row=self.iCurr, column=6).value = companyIncome
        initAcct.insProfits = toNumberFmt(initAcct.insProfits + insProfits)
        self.ws.cell(row=self.iCurr, column=7).value = insProfits
        initAcct.chnlAmt = toNumberFmt(initAcct.chnlAmt + chnlAmt)
        self.ws.cell(row=self.iCurr, column=8).value = chnlAmt
        self.iCurr = self.iCurr + 1

    #����-�쳣��δʶ���̻�
    def recordUnknownMchtTxn(self):
        self.ws.cell(row=self.iCurr, column=2).value = '����-�쳣��δʶ���̻�'
        self.iCurr = self.iCurr + 1

    #���� - �쳣������ʶ���̻�
    def recordAbnormalWriteOff(self):
        self.ws.cell(row=self.iCurr, column=2).value = '���� - �쳣������ʶ���̻�'
        self.iCurr = self.iCurr + 1

    # ����
    def recordMchtStlmAmtOutS0(self, initAcct, outMchtPayAmt, othLoan, sandLoan):
        self.ws.cell(row=self.iCurr, column=2).value = '����'
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt - outMchtPayAmt)
        self.ws.cell(row=self.iCurr, column=3).value = toNumberFmt(0 - outMchtPayAmt)
        initAcct.payChnlLoan = toNumberFmt(initAcct.payChnlLoan + othLoan)
        self.ws.cell(row=self.iCurr, column=9).value = othLoan
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit + sandLoan)
        self.ws.cell(row=self.iCurr, column=10).value = sandLoan
        self.iCurr = self.iCurr + 1

    #�����˵�,����ֵΪ��
    def recordChargeBack(self, initAcct, amtBack):
        self.ws.cell(row=self.iCurr, column=2).value = '�����˵�'
        initAcct.riskLoan = toNumberFmt(initAcct.riskLoan + amtBack)
        self.ws.cell(row=self.iCurr, column=5).value = amtBack
        initAcct.chnlAmt = toNumberFmt(initAcct.chnlAmt - amtBack)
        self.ws.cell(row=self.iCurr, column=8).value = toNumberFmt(0 - amtBack)
        self.iCurr = self.iCurr + 1

    #�����˵����������
    def recordChargeBackOff(self, initAcct, amtBackOff):
        self.ws.cell(row=self.iCurr, column=2).value = '�����˵����������'
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt - amtBackOff)
        self.ws.cell(row=self.iCurr, column=3).value = toNumberFmt(0 - amtBackOff)
        initAcct.riskLoan = toNumberFmt(initAcct.riskLoan + amtBackOff)
        self.ws.cell(row=self.iCurr, column=5).value = amtBackOff
        self.iCurr = self.iCurr + 1

    #����-������������δ��(��������ʱ����)
    def recordChnlUnstlm(self):
        self.ws.cell(row=self.iCurr, column=2).value = '����-������������δ��'
        self.iCurr = self.iCurr + 1

    #�̻����ֹ��˻���֤��򷢿��˵��ʽ�
    def recordMchtDeposit(self, initAcct, amtDeposit):
        self.ws.cell(row=self.iCurr, column=2).value = '�̻����ֹ��˻���֤��򷢿��˵��ʽ�'
        self.ws.cell(row=self.iCurr, column=2).alignment = Alignment(wrap_text=True)
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt + amtDeposit)
        self.ws.cell(row=self.iCurr, column=3).value = amtDeposit
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit - amtDeposit)
        self.ws.cell(row=self.iCurr, column=10).value = toNumberFmt(0 - amtDeposit)
        self.iCurr = self.iCurr + 1

    #�̻����ֹ��˻���֤��򷢿��˵��ʽ𣩹���
    def recordMchtDepositHanging(self):
        self.ws.cell(row=self.iCurr, column=2).value = '�̻����ֹ��˻���֤��򷢿��˵��ʽ𣩹���'
        self.ws.cell(row=self.iCurr, column=2).alignment = Alignment(wrap_text=True)
        self.iCurr = self.iCurr + 1

    #�̻����ֹ��˻���֤��򷢿��˵��ʽ𣩹���ȷ��
    def recordMDepConfirm(self):
        self.ws.cell(row=self.iCurr, column=2).value = '�̻����ֹ��˻���֤��򷢿��˵��ʽ𣩹���ȷ��'
        self.ws.cell(row=self.iCurr, column=2).alignment = Alignment(wrap_text=True)
        self.iCurr = self.iCurr + 1

    #�̻�����˻أ��ֹ��˻�������
    def recordMchtDepositReturn(self):
        self.ws.cell(row=self.iCurr, column=2).value = '�̻�����˻أ��ֹ��˻�������'
        self.iCurr = self.iCurr + 1

    # ��ط��𶳽�
    def recordLockTxn(self, initAcct, lockAmt):
        self.ws.cell(row=self.iCurr, column=2).value = '��ط��𶳽�'
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt - lockAmt)
        self.ws.cell(row=self.iCurr, column=3).value = toNumberFmt(0 - lockAmt)
        initAcct.lockAmt = toNumberFmt(initAcct.lockAmt + lockAmt)
        self.ws.cell(row=self.iCurr, column=12).value = lockAmt
        self.iCurr = self.iCurr + 1

    #��ط���ⶳ
    def recordUnlockTxn(self, initAcct, unlockAmt):
        self.ws.cell(row=self.iCurr, column=2).value = '��ط���ⶳ'
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt + unlockAmt)
        self.ws.cell(row=self.iCurr, column=3).value = unlockAmt
        initAcct.lockAmt = toNumberFmt(initAcct.lockAmt - unlockAmt)
        self.ws.cell(row=self.iCurr, column=12).value = toNumberFmt(0 - unlockAmt)
        self.iCurr = self.iCurr + 1

    #���ʷ���
    def recordOthLoan(self):
        self.ws.cell(row=self.iCurr, column=2).value = '���ʷ���'
        self.iCurr = self.iCurr + 1

    #����׷��
    def recordOthLoanOff(self):
        self.ws.cell(row=self.iCurr, column=2).value = '����׷��'
        self.iCurr = self.iCurr + 1

    # ���
    def recordInAmt(self, initAcct, chnlAmt):
        self.ws.cell(row=self.iCurr, column=2).value = '���'
        initAcct.chnlAmt = toNumberFmt(initAcct.chnlAmt + chnlAmt)
        self.ws.cell(row=self.iCurr, column=8).value = chnlAmt
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit - chnlAmt)
        self.ws.cell(row=self.iCurr, column=10).value = toNumberFmt(0 - chnlAmt)
        self.iCurr = self.iCurr + 1

    #�̻���������,T1������
    def recordMchtStlmAmtOutT1(self, initAcct, outMchtPayAmt, othLoan, sandLoan):
        self.ws.cell(row=self.iCurr, column=2).value = '�̻���������'
        initAcct.mchtStlmAmt = toNumberFmt(initAcct.mchtStlmAmt - outMchtPayAmt)
        self.ws.cell(row=self.iCurr, column=3).value = toNumberFmt(0 - outMchtPayAmt)
        initAcct.payChnlLoan = toNumberFmt(initAcct.payChnlLoan + othLoan)
        self.ws.cell(row=self.iCurr, column=9).value = othLoan
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit + sandLoan)
        self.ws.cell(row=self.iCurr, column=10).value = sandLoan
        self.iCurr = self.iCurr + 1


    #��˾�����
    def recordCompanyIncomePayOut(self, initAcct, outAmt):
        self.ws.cell(row=self.iCurr, column=2).value = '��˾�����'
        initAcct.companyIncome = toNumberFmt(initAcct.companyIncome - outAmt)
        self.ws.cell(row=self.iCurr, column=6).value = toNumberFmt(0 - outAmt)
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit + outAmt)
        self.ws.cell(row=self.iCurr, column=10).value = outAmt
        self.iCurr = self.iCurr + 1

    #�����������
    def recordInsIncomePayOut(self, initAcct, outAmt, sandLoan):
        self.ws.cell(row=self.iCurr, column=2).value = '�����������'
        initAcct.insProfits = toNumberFmt(initAcct.insProfits - outAmt)
        self.ws.cell(row=self.iCurr, column=7).value = toNumberFmt(0 - outAmt)
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit + sandLoan)
        self.ws.cell(row=self.iCurr, column=10).value = sandLoan
        self.iCurr = self.iCurr + 1

    #�ͽ��˻�-�Թ������Ĵ����˵�
    def recordPayReturnPublic(self):
        self.ws.cell(row=self.iCurr, column=2).value = '�ͽ��˻�'
        self.iCurr = self.iCurr + 1
    #���������˻�
    def recordPayReturnPrivate(self):
        self.ws.cell(row=self.iCurr, column=2).value = '���������˻�'
        self.iCurr = self.iCurr + 1

    #�ʽ������۴���
    def recordChnlPayAmt(self, initAcct, chnlPayAmt):
        self.ws.cell(row=self.iCurr, column=2).value = '�ʽ������۴���'
        initAcct.payChnlLoan = toNumberFmt(initAcct.payChnlLoan - chnlPayAmt)
        self.ws.cell(row=self.iCurr, column=9).value = toNumberFmt(0 - chnlPayAmt)
        initAcct.bankDeposit = toNumberFmt(initAcct.bankDeposit + chnlPayAmt)
        self.ws.cell(row=self.iCurr, column=10).value = chnlPayAmt
        self.iCurr = self.iCurr + 1

    def recordFinalAmt(self, initAcct):
        self.__recordDtlAmt(initAcct, '��ĩ')

    def __recordDtlAmt(self, initAcct, item):
        self.ws.cell(row=self.iCurr, column=1).value = item
        self.ws.cell(row=self.iCurr, column=3).value = initAcct.mchtStlmAmt
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.abnormalDeposit
        self.ws.cell(row=self.iCurr, column=5).value = initAcct.riskLoan
        self.ws.cell(row=self.iCurr, column=6).value = initAcct.companyIncome
        self.ws.cell(row=self.iCurr, column=7).value = initAcct.insProfits
        self.ws.cell(row=self.iCurr, column=8).value = initAcct.chnlAmt
        self.ws.cell(row=self.iCurr, column=9).value = initAcct.payChnlLoan
        self.ws.cell(row=self.iCurr, column=10).value = initAcct.bankDeposit
        self.ws.cell(row=self.iCurr, column=11).value = initAcct.diffAmt
        self.ws.cell(row=self.iCurr, column=12).value = initAcct.riskLoan
        self.ws.cell(row=self.iCurr, column=13).value = initAcct.othLoan
        self.iCurr = self.iCurr + 1

    #�б����
    def recordCheckCol(self, initAcct):
        self.ws.cell(row=self.iCurr, column=1).value = '�յ�����ƽ̨'
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '�̻��������'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.mchtStlmAmt
        amt = toNumberFmt(initAcct.mchtStlmAmt)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '�����쳣����'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.abnormalDeposit
        amt = toNumberFmt(amt + initAcct.abnormalDeposit)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '���յ���'
        self.ws.cell(row=self.iCurr, column=4).value = toNumberFmt(0 - initAcct.riskLoan)
        amt = toNumberFmt(amt + initAcct.riskLoan)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = 'δ����˾����'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.companyIncome
        amt = toNumberFmt(amt + initAcct.companyIncome)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = 'δ������������'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.insProfits
        amt = toNumberFmt(amt + initAcct.insProfits)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = 'Ӧ������'
        self.ws.cell(row=self.iCurr, column=4).value = toNumberFmt(0 - initAcct.chnlAmt)
        amt = toNumberFmt(amt + initAcct.chnlAmt)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = 'Ӧ��������������'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.payChnlLoan
        amt = toNumberFmt(amt + initAcct.payChnlLoan)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '���д��'
        self.ws.cell(row=self.iCurr, column=4).value = toNumberFmt(0 - initAcct.bankDeposit)
        amt = toNumberFmt(amt + initAcct.bankDeposit)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '����δ����'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.diffAmt
        amt = toNumberFmt(amt + initAcct.diffAmt)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '��ض���'
        self.ws.cell(row=self.iCurr, column=4).value = initAcct.riskLoan
        amt = toNumberFmt(amt + initAcct.riskLoan)
        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=2).value = '��'
        self.ws.cell(row=self.iCurr, column=3).value = '��������'
        self.ws.cell(row=self.iCurr, column=4).value = toNumberFmt(0 - initAcct.othLoan)
        amt = toNumberFmt(amt + initAcct.othLoan)
        self.iCurr = self.iCurr + 1

        self.iCurr = self.iCurr + 1
        self.ws.cell(row=self.iCurr, column=3).value = '�ʽ�У��'
        self.ws.cell(row=self.iCurr, column=4).value = amt
        col = self.ws.column_dimensions['C']
        col.width = 17.0
        col = self.ws.column_dimensions['A']
        col.width = 15.0


class stlmPvsnAcctInfo:
    def __init__(self, db, stlmDate):
        self.db = db
        self.stlmDate = stlmDate
        self.mchtStlmAmt = 0
        self.companyIncome = 0
        self.insProfits = 0
        self.chnlAmt = 0
        self.diffAmt = 0
        self.abnormalDeposit = 0
        self.lockAmt = 0
        self.bankDeposit = 0
        self.riskLoan = 0
        self.othLoan = 0
        self.payChnlLoan = 0
        self.__get_init_acct_info()

    #��ȡ������ĩ
    def __get_init_acct_info(self):
        sql = "select MCHT_STLM_AMT,COMPANY_INCOME,INS_PROFITS,CHNL_AMT, " \
              "DIFF_AMT,ABNORML_DEPOSIT,LOCK_AMT,BANK_DEPOSIT, " \
              "RISK_LOAN,OTH_LOAN,PAY_CHNL_LOAN  from TBL_STLM_PVSN_RPT where host_date ='%s'" % getLastDay(self.stlmDate)
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            self.mchtStlmAmt = toNumberFmt(x[0])
            self.companyIncome = toNumberFmt(x[1])
            self.insProfits = toNumberFmt(x[2])
            self.chnlAmt = toNumberFmt(x[3])
            self.diffAmt = toNumberFmt(x[4])
            self.abnormalDeposit = toNumberFmt(x[5])
            self.lockAmt = toNumberFmt(x[6])
            self.bankDeposit = toNumberFmt(x[7])
            self.riskLoan = toNumberFmt(x[8])
            self.othLoan = toNumberFmt(x[9])
            self.payChnlLoan = toNumberFmt(x[10])

    #��ֵ��¼���ݿ�
    def insertFinalInfo(self):
        sql = "insert into TBL_STLM_PVSN_RPT "
        sqltmp = "(host_date,MCHT_STLM_AMT,COMPANY_INCOME,INS_PROFITS,CHNL_AMT, " \
                 "DIFF_AMT,ABNORML_DEPOSIT,LOCK_AMT,BANK_DEPOSIT, " \
                 "RISK_LOAN,OTH_LOAN,PAY_CHNL_LOAN ) values "
        sql = sql + sqltmp
        sqltmp = "('%s',%f, %f, %f, %f, %f, %f, %f," \
                 "%f, %f, %f, %f)" % \
                 (self.stlmDate, self.mchtStlmAmt, self.companyIncome, self.insProfits, self.chnlAmt,
                  self.diffAmt, self.abnormalDeposit, self.lockAmt, self.bankDeposit, self.riskLoan,
                  self.othLoan, self.payChnlLoan)
        sql = sql + sqltmp
        cursor = self.db.cursor()
        print(sql)
        cursor.execute(sql)
        self.db.commit()
        cursor.close()

#���������ֵͳ��
class txnInfo:
    def __init__(self, db, stlmDate):
        self.db = db
        self.stlmDate = stlmDate

        self.__get_mcht_stlm()
        self.__get_company_income()
        self.__get_ins_income()
        self.__get_risk_loan()
        self.__get_paytxn_amt()

    #����������
    def __get_mcht_stlm(self):
        sql = "select sum(TRANS_AMT - TRANS_FEE) from TBL_INS_PROFITS_TXN_SUM where " \
              "host_date ='%s' and PROFITS_TYPE ='00'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.mchtStlmAmt = toNumberFmt(x[0])
        cursor.close()

    #�����쳣�������,���ճ������������
    def __get_last_day_off_txn(self):
        sql = "select "


    #��˾δ������(��˾����),��˾�������
    def __get_company_income(self):
        sql = "select sum(ALL_PROFITS) from TBL_SAND_ACQ_PROFITS where " \
              "host_date ='%s'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.companyIncome = toNumberFmt(x[0])
        sql = "select sum(PROFITS_AMT) from TBL_SAND_PROFITS_CHARGE_OFFS where CHARGE_DATE = '%s'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.companyIncomeOff = toNumberFmt(x[0])
        cursor.close()

    #��������������
    def __get_ins_income(self):
        sql = "select sum(ALL_PROFITS) from TBL_INS_PROFITS_TXN_SUM where " \
              "host_date ='%s'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.insIncome = toNumberFmt(x[0])
        cursor.close()


    #���յ���
    def __get_risk_loan(self):
        #��ʱ��֪����������,��ʱΪ��
        self.riskLoan = 0

    #�ɹ��������,����S0,T1����
    def __get_paytxn_amt(self):
        #S0
        sql = "select sum(REAL_TRANS_AMT) from tbl_stlm_txn_bill_dtl where " \
              "chnl_id ='A002' and host_date ='%s' and pay_type = '00'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.mchtPayAmtS0 = toNumberFmt(0 - x[0])

        #T1
        sql = "select sum(REAL_TRANS_AMT) from tbl_stlm_txn_bill_dtl where " \
              "chnl_id ='A002' and host_date ='%s' and pay_type != '00'" % self.stlmDate
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.mchtPayAmtT1 = toNumberFmt(0 - x[0])

        #����������
        sql = "select sum(trans_amt/100) from tbl_acq_txn_log where host_date ='%s' and " \
              "txn_num ='1801' and substrb(ADDTNL_DATA,1,2) ='04' and trans_state ='1'" % self.stlmDate
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.agentPayAmt = toNumberFmt(x[0])
            self.mchtPayAmtT1 = toNumberFmt(self.mchtPayAmtT1 - self.agentPayAmt)
        cursor.close()



class lockInfo:
    def __init__(self, db, stlmDate):
        self.db = db
        self.stlmDate = stlmDate
        self.lockAmt = 0
        self.unlockAmt = 0

        self.__get_lock_amt()

    def __get_lock_amt(self):
        sql = "select sum(LOCK_AT)/100 from T_TXN_LOCK where " \
              "host_date ='%s' and TXN_TYPE ='01' " % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.lockAmt = toNumberFmt(x[0])
        cursor.close()

    def __get_unlock_amt(self):
        sql = "select sum(free_at)/100 from T_TXN_LOCK where" \
              "host_date ='%s' and TXN_TYPE ='02' " % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.unlockAmt = toNumberFmt(x[0])
        cursor.close()

class chnlAmtInfo:
    def __init__(self, db, stlmDate):
        self.db = db
        self.stlmDate = stlmDate

        self.__get_intxn_amt()

    #�������ļ�,������Ҫ�ڼ��ջ���
    def __get_intxn_amt(self):
        self.intxnAmt = 0
        #���ڼ���
        if isHoliDay(self.db, self.stlmDate):
            return

        #�ڼ��ս���
        cursor = self.db.cursor()
        start_date = getLastDay(self.stlmDate)
        end_date = start_date
        sql = "select START_DATE,END_DATE from TBL_HOLI_INF where END_DATE ='%s'" % self.stlmDate
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            start_date = getLastDay(x[0])

        sql = "select sum(REAL_TRANS_AMT - iss_fee - swt_fee - prod_fee) from tbl_stlm_txn_bill_dtl where chnl_id ='A001' and " \
              "stlm_date >='%s' and stlm_date <='%s'" % (start_date, end_date)
        cursor.execute(sql)
        x = cursor.fetchone()
        if x[0] is not None:
            self.intxnAmt = toNumberFmt(x[0])
        cursor.close()

class chnlPayAmtInfo:
    def __init__(self, db, stlmDate):
        self.db = db
        self.stlmDate = stlmDate
        self.__get_chnl_loan_amt()
        self.__get_pay_txn_amt()

    #���������ļ������ۿ�(�ڼ��ռ���)
    def __get_chnl_loan_amt(self):
        self.othAllLoan = 0.0
        if isHoliDay(self.db, self.stlmDate):
            return

        #�ǽڼ��ջ�ȡǰһ��,��ڼ�������ĵ��ʷ���
        cursor = self.db.cursor()
        start_date = getLastDay(self.stlmDate)
        end_date = start_date
        sql = "select START_DATE,END_DATE from TBL_HOLI_INF where END_DATE ='%s'" % self.stlmDate
        cursor.execute(sql)
        x = cursor.fetchone()
        if x is not None:
            start_date = getLastDay(x[0])

        sql = "select nvl(sum(REAL_TRANS_AMT),0), DEST_CHNL_ID from TBL_STLM_TXN_BILL_DTL where stlm_date >='%s' " \
              " and stlm_date <='%s' group by DEST_CHNL_ID" % (start_date, end_date)
        cursor.execute(sql)
        for ltData in cursor:
            #ͨ������
            if getFundType(ltData[1]) == 0:
                self.othAllLoan = toNumberFmt(self.othAllLoan - ltData[0])


    #���˺��׷���
    def __get_pay_txn_amt(self):
        self.sandLoanT1 = 0.0
        self.sandLoanS0 = 0.0
        self.othLoanT1 = 0.0
        self.othLoanS0 = 0.0
        self.agentIncomePay = 0.0
        sql = "select nvl(sum(REAL_TRANS_AMT),0), DEST_CHNL_ID, pay_type from TBL_STLM_TXN_BILL_DTL where stlm_date ='%s' " \
              " group by DEST_CHNL_ID,pay_type" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        for ltData in cursor:
            #��ȡ������������
            if getFundType(ltData[1]) == 1:
                #ɼ�µ���
                if str(ltData[1]).rstrip() == '00000910':
                    self.agentIncomePay = toNumberFmt(self.agentIncomePay - ltData[0])
                elif ltData[2] == '00':
                    self.sandLoanS0 = toNumberFmt(self.sandLoanS0 - ltData[0])
                else:
                    self.sandLoanT0 = toNumberFmt(self.sandLoanT0 - ltData[0])
            elif getFundType(ltData[1]) == 0:
                if ltData[2] == '00':
                    self.othLoanS0 = toNumberFmt(self.othLoanS0 - ltData[0])
                else:
                    self.othLoanT1 = toNumberFmt(self.othLoanT1 - ltData[0])

        cursor.close()



def main():
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),
                           encoding='gb18030')
    dbacc = cx_Oracle.connect('%s/%s@%s' % (os.environ['ACCDBUSER'], os.environ['ACCDBPWD'], os.environ['TNSNAME']),
                           encoding='gb18030')
    # ��ȡ������
    if len(sys.argv) == 1:
        cursor = db.cursor()
        sql = "select BF_STLM_DATE from TBL_BAT_CUT_CTL"
        cursor.execute(sql)
        x = cursor.fetchone()
        stlm_date = x[0]
        cursor.close()
    else:
        stlm_date = sys.argv[1]
    print('hostDate %s genStlmPvsnRpt begin' % stlm_date)

    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_date)
    filename = filePath + 'StlmPvsnRpt_%s.xlsx' % stlm_date

    stlmPvsnAcct = stlmPvsnAcctInfo(db, stlm_date)
    wb = Workbook()
    ws = wb.active
    rptxls = rptFile(ws)
    rptxls.setTitle('��ϸ����')
    rptxls.head()
    rptxls.recordInitAmt(stlmPvsnAcct)

    #���콻��
    txnInf = txnInfo(db, stlm_date)
    rptxls.recordTodayTxn(stlmPvsnAcct, txnInf.mchtStlmAmt, txnInf.companyIncome,
                          txnInf.insIncome, txnInf.calcChnlAmt)
    #����-�쳣��δʶ���̻�
    rptxls.recordUnknownMchtTxn()

    #���� - �쳣������ʶ���̻�
    rptxls.recordAbnormalWriteOff()
    chnlPayAmt = chnlPayAmtInfo(db, stlm_date)

    #����
    rptxls.recordMchtStlmAmtOutS0(stlmPvsnAcct, txnInf.mchtPayAmtS0, chnlPayAmt.othLoanS0, chnlPayAmt.sandLoanS0)

    # �����˵�,����ֵΪ��
    rptxls.recordChargeBack(stlmPvsnAcct, 0)

    #�����˵����������
    rptxls.recordChargeBackOff(stlmPvsnAcct, 0)

    #����-������������δ��(��������ʱ����)
    rptxls.recordChnlUnstlm()

    #�̻����ֹ��˻���֤��򷢿��˵��ʽ�
    rptxls.recordMchtDeposit(stlmPvsnAcct, 0)

    #�̻����ֹ��˻���֤��򷢿��˵��ʽ𣩹���
    rptxls.recordMchtDepositHanging()

    #�̻����ֹ��˻���֤��򷢿��˵��ʽ𣩹���ȷ��
    rptxls.recordMDepConfirm()

    #�̻�����˻أ��ֹ��˻�������
    rptxls.recordMchtDepositReturn()

    #��ط��𶳽�
    lockInf = lockInfo(dbacc, stlm_date)
    rptxls.recordLockTxn(stlmPvsnAcct, lockInf.lockAmt)
    #��ط���ⶳ
    rptxls.recordUnlockTxn(stlmPvsnAcct, lockInf.unlockAmt)

    #���ʷ���
    rptxls.recordOthLoan()

    #����׷��
    rptxls.recordOthLoanOff()

    #���
    chnlFile = chnlAmtInfo(db, stlm_date)
    rptxls.recordInAmt(stlmPvsnAcct, chnlFile.intxnAmt)
    #�̻���������,T1������
    rptxls.recordMchtStlmAmtOutT1(stlmPvsnAcct, txnInf.mchtPayAmtT1, chnlPayAmt.othLoanT1, chnlPayAmt.sandLoanT1)

    #��˾�����
    rptxls.recordCompanyIncomePayOut(stlmPvsnAcct, txnInf.companyIncomeOff)

    #�����������
    rptxls.recordInsIncomePayOut(stlmPvsnAcct, txnInf.agentPayAmt, chnlPayAmt.agentIncomePay)

    #�ͽ��˻�-�Թ������Ĵ����˵�
    rptxls.recordPayReturnPublic()

    #���������˻�
    rptxls.recordPayReturnPrivate()

    #�ʽ������۴���
    rptxls.recordChnlPayAmt(stlmPvsnAcct, chnlPayAmt.othAllLoan)

    #��ĩ
    rptxls.recordFinalAmt(stlmPvsnAcct)

    col = ws.column_dimensions['B']
    col.width = 30.0

    rptxls2 = rptFile(wb.create_sheet())
    rptxls2.setTitle('�б�У��')
    rptxls2.recordCheckCol(stlmPvsnAcct)

    wb.save(filename)
    wb.close()
    stlmPvsnAcct.insertFinalInfo()
    print('hostDate %s genStlmPvsnRpt end' % stlm_date)

if __name__ == '__main__':
    main()