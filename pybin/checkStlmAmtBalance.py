#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#����ƽ���,�˶������ʽ��������ʽ��ϵ

import os
import sys
import cx_Oracle
from openpyxl.workbook import Workbook
from utl.common import *

#ͨ���ļ���Ϣ
class chnlBillClass:
    def __init__(self, db, stlmDate):
        self.db = db
        self.stlmDate = stlmDate
        self.__calcAllAmt()
        self.__calcErrAmt()

    def __calcAllAmt(self):
        sql = "select count(*),sum(REAL_TRANS_AMT),sum(iss_fee), sum(swt_fee), sum(prod_fee) from TBL_STLM_TXN_BILL_DTL where stlm_date ='%s' and chnl_id ='A001'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        cursor.close()
        if x is not None:
            self.txnCount = toNumberFmt(x[0])
            self.txnAmt = toNumberFmt(x[1])
            self.issFee = toNumberFmt(x[2])
            self.swtFee = toNumberFmt(x[3])
            self.prodFee = toNumberFmt(x[4])
        else:
            self.txnCount = 0
            self.txnAmt = 0
            self.issFee = 0
            self.swtFee = 0
            self.prodFee = 0
        self.allCost = toNumberFmt(self.issFee + self.swtFee + self.prodFee)
        self.stlmAmt = toNumberFmt(self.txnAmt - self.allCost)

    # ���㵱�ղ���ʽ�
    def __calcErrAmt(self):
        sql = "select sum(REAL_TRANS_AMT) from tbl_stlm_txn_bill_dtl where stlm_date ='%s' and " \
              "txn_num !='1011' and chnl_id ='A001'" % self.stlmDate
        cursor = self.db.cursor()
        cursor.execute(sql)
        x = cursor.fetchone()
        cursor.close()
        if x is not None:
            self.errAmt = toNumberFmt(x[0])
        else:
            self.errAmt = 0

#���㱾������������
def calcStlmInnerAmt(db, stlm_date):
    sql = "select sum(trans_amt)/100 from tbl_acq_txn_log where host_date = '%s' and txn_num ='1011' " \
          "and trans_state ='1' and REVSAL_FLAG ='0' and CANCEL_FLAG ='0'" % stlm_date
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    if x is not None:
        return toNumberFmt(x[0])
    else:
        return 0

#�������ճ����ʽ�
def calcLastChnlStlmFunds(db, stlm_date):
    sql = "select sum(REAL_TRANS_AMT) from tbl_stlm_txn_bill_dtl where host_date ='%s' and " \
          "stlm_date = '%s' and check_sta ='1' and txn_num ='1011'"\
          % (stlm_date, getLastDay(stlm_date))
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    if x is not None:
        return toNumberFmt(x[0])
    else:
        return 0

#���㱾�ճ����ʽ�
def calcChnlStlmFunds(db, stlm_date):
    #ͨ���ļ�����,��˾�����մ���
    sql = "select sum(REAL_TRANS_AMT) from tbl_stlm_txn_bill_dtl where host_date ='%s' and " \
          "stlm_date = '%s' and check_sta ='1' and txn_num ='1011'" \
          % (getNextDay(stlm_date), stlm_date)
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    if x is not None:
        return toNumberFmt(x[0])
    else:
        return 0

#������˳���׽��
def calcLongTxnAmt(db, stlm_date):
    sql = "select sum(CHNL_TXN_AMT) from tbl_err_chk_txn_dtl where " \
          "host_date = '%s' and CHK_STA ='1' and group_id ='A001'" % stlm_date
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    cursor.close()
    if x is not None:
        return toNumberFmt(x[0])
    else:
        return 0


def main():
    # ���ݿ���������
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),encoding='gb18030')
    # ��ȡ������
    if len(sys.argv) == 1:
        cursor = db.cursor()
        sql = "select BF_STLM_DATE from TBL_BAT_CUT_CTL"
        cursor.execute(sql)
        x = cursor.fetchone()
        stlm_date = x[0]
        cursor.close()
    else:
        stlm_date = sys.argv[1]
    print('hostDate %s rpt begin' % stlm_date)

    chnlbill = chnlBillClass(db, stlm_date)
    innerAmt = calcStlmInnerAmt(db, stlm_date)
    lastChnlFunds = calcLastChnlStlmFunds(db, stlm_date)
    chnlFunds = calcChnlStlmFunds(db, stlm_date)
    longAmt = calcLongTxnAmt(db, stlm_date)

    print("innerAmt:%.2f" % innerAmt)
    print("lastChnlFunds:%.2f" % lastChnlFunds)
    print("chnlFunds:%.2f" % chnlFunds)
    print("longAmt:%.2f" % longAmt)
    print("errAmt:%.2f" % chnlbill.errAmt)
    print("chnlAmt:%.2f" % chnlbill.txnAmt)

    if toNumberFmt(innerAmt - lastChnlFunds + chnlFunds + chnlbill.errAmt + longAmt) != toNumberFmt(chnlbill.txnAmt):
        bal_sta = '2'
    else:
        bal_sta = '1'


    #���������Ҫ�����ļ�
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=8).value = '�յ�����ϵͳ�����ʽ�ƽ���'
    i = 2
    ws.cell(row=i, column=1).value = '���������ܽ��'
    ws.cell(row=i, column=2).value = '���ճ����ʽ�'
    ws.cell(row=i, column=3).value = '���ճ����ʽ�'
    ws.cell(row=i, column=4).value = '����ཻ�׽��'
    ws.cell(row=i, column=5).value = '�����׽��'
    ws.cell(row=i, column=6).value = 'ͨ���ļ����'
    ws.cell(row=i, column=7).value = '�˶Խ��'
    i = i + 1
    ws.cell(row=i, column=1).value = innerAmt
    ws.cell(row=i, column=2).value = lastChnlFunds
    ws.cell(row=i, column=3).value = chnlFunds
    ws.cell(row=i, column=4).value = chnlbill.errAmt
    ws.cell(row=i, column=5).value = longAmt
    ws.cell(row=i, column=6).value = chnlbill.txnAmt
    if bal_sta == '1':
        ws.cell(row=i, column=7).value = 'ƽ��'
    else:
        ws.cell(row=i, column=7).value = '��ƽ'
    i = i + 2

    #ͨ���ļ�����
    ws.cell(row=i, column=1).value = 'ͨ���ļ��ܶ�'
    i = i + 1
    ws.cell(row=i, column=1).value = '���ױ���'
    ws.cell(row=i, column=2).value = '���ױ���'
    ws.cell(row=i, column=3).value = '���������'
    ws.cell(row=i, column=4).value = '��������ת�ӷ�'
    ws.cell(row=i, column=5).value = 'Ʒ�Ʒ����'
    ws.cell(row=i, column=6).value = '�ܳɱ�'
    ws.cell(row=i, column=7).value = '�ʽ����㾻��'
    i = i + 1
    ws.cell(row=i, column=1).value = chnlbill.txnAmt
    ws.cell(row=i, column=2).value = chnlbill.txnCount
    ws.cell(row=i, column=3).value = chnlbill.issFee
    ws.cell(row=i, column=4).value = chnlbill.swtFee
    ws.cell(row=i, column=5).value = chnlbill.prodFee
    ws.cell(row=i, column=6).value = chnlbill.allCost
    ws.cell(row=i, column=7).value = chnlbill.stlmAmt
    i = i + 1

    #���ճ�������ϸ������
    if chnlFunds > 0:
        ws.cell(row=i, column=1).value = '���ճ�������ϸ������'
        i = i + 1
        ws.cell(row=i, column=1).value = '��������'
        ws.cell(row=i, column=2).value = 'ʱ��'
        ws.cell(row=i, column=3).value = '�̻���'
        ws.cell(row=i, column=4).value = '����'
        ws.cell(row=i, column=5).value = '���'
        ws.cell(row=i, column=6).value = '���������'
        ws.cell(row=i, column=7).value = '��������ת�ӷ�'
        ws.cell(row=i, column=8).value = 'Ʒ�Ʒ����'
        ws.cell(row=i, column=9).value = '�ܳɱ�'
        ws.cell(row=i, column=10).value = '�ʽ����㾻��'
        i = i + 1

        sql = "select TXN_POS_DATE, TXN_POS_TIME, mcht_cd, pan, REAL_TRANS_AMT, iss_fee, swt_fee, prod_fee from " \
              " tbl_stlm_txn_bill_dtl where host_date ='%s' and " \
          "stlm_date = '%s' and check_sta ='1' and txn_num ='1011'" \
          % (getNextDay(stlm_date), stlm_date)
        cursor = db.cursor()
        cursor.execute(sql)
        allTxnAmt = 0
        allIssAmt = 0
        allSwtAmt = 0
        allProdAmt = 0
        allCost = 0
        allStlmAmt = 0
        for ltTxn in cursor:
            ws.cell(row=i, column=1).value = ltTxn[0]
            ws.cell(row=i, column=2).value = ltTxn[1]
            ws.cell(row=i, column=3).value = ltTxn[2]
            ws.cell(row=i, column=4).value = ltTxn[3]
            allTxnAmt = allTxnAmt + toNumberFmt(ltTxn[4])
            ws.cell(row=i, column=5).value = toNumberFmt(ltTxn[4])
            allIssAmt = allIssAmt + toNumberFmt(ltTxn[5])
            ws.cell(row=i, column=6).value = toNumberFmt(ltTxn[5])
            allSwtAmt = allSwtAmt + toNumberFmt(ltTxn[6])
            ws.cell(row=i, column=7).value = toNumberFmt(ltTxn[6])
            allProdAmt = allProdAmt + toNumberFmt(ltTxn[7])
            ws.cell(row=i, column=8).value = toNumberFmt(ltTxn[7])
            allCost = allCost + toNumberFmt(ltTxn[5]) + toNumberFmt(ltTxn[6]) + toNumberFmt(ltTxn[7])
            ws.cell(row=i, column=9).value = toNumberFmt(ltTxn[5]) + toNumberFmt(ltTxn[6]) + toNumberFmt(ltTxn[7])
            allStlmAmt = allStlmAmt + toNumberFmt(ltTxn[4]) - toNumberFmt(ltTxn[5]) + toNumberFmt(ltTxn[6]) + toNumberFmt(ltTxn[7])
            ws.cell(row=i, column=10).value = toNumberFmt(ltTxn[4]) - toNumberFmt(ltTxn[5]) + toNumberFmt(ltTxn[6]) + toNumberFmt(ltTxn[7])
            i = i + 1
        #�ܼ�
        ws.cell(row=i, column=1).value = '�ܼ�:'
        ws.cell(row=i, column=5).value = allTxnAmt
        ws.cell(row=i, column=6).value = allIssAmt
        ws.cell(row=i, column=7).value = allSwtAmt
        ws.cell(row=i, column=8).value = allProdAmt
        ws.cell(row=i, column=9).value = allCost
        ws.cell(row=i, column=10).value = allStlmAmt

        # �������ϸ������
        if chnlbill.errAmt > 0:
            ws.cell(row=i, column=1).value = '�������ϸ������'
            i = i + 1
            ws.cell(row=i, column=1).value = '��������'
            ws.cell(row=i, column=2).value = 'ʱ��'
            ws.cell(row=i, column=3).value = '�̻���'
            ws.cell(row=i, column=4).value = '����'
            ws.cell(row=i, column=5).value = '���'
            ws.cell(row=i, column=6).value = '���������'
            ws.cell(row=i, column=7).value = '��������ת�ӷ�'
            ws.cell(row=i, column=8).value = 'Ʒ�Ʒ����'
            ws.cell(row=i, column=9).value = '����'
            ws.cell(row=i, column=10).value = '�ܳɱ�'
            ws.cell(row=i, column=11).value = '�ʽ����㾻��'
            i = i + 1

            sql = "select TXN_POS_DATE, TXN_POS_TIME, mcht_cd, pan, REAL_TRANS_AMT, iss_fee, swt_fee, prod_fee, err_fee from " \
                  " tbl_stlm_txn_bill_dtl where " \
                  "stlm_date = '%s' and txn_num in ('9009','9005')" \
                  % (stlm_date)
            cursor = db.cursor()
            cursor.execute(sql)
            allTxnAmt = 0
            allIssAmt = 0
            allSwtAmt = 0
            allProdAmt = 0
            allErrFee = 0
            allCost = 0
            allStlmAmt = 0
            for ltTxn in cursor:
                ws.cell(row=i, column=1).value = ltTxn[0]
                ws.cell(row=i, column=2).value = ltTxn[1]
                ws.cell(row=i, column=3).value = ltTxn[2]
                ws.cell(row=i, column=4).value = ltTxn[3]
                allTxnAmt = allTxnAmt + toNumberFmt(ltTxn[4])
                ws.cell(row=i, column=5).value = toNumberFmt(ltTxn[4])
                allIssAmt = allIssAmt + toNumberFmt(ltTxn[5])
                ws.cell(row=i, column=6).value = toNumberFmt(ltTxn[5])
                allSwtAmt = allSwtAmt + toNumberFmt(ltTxn[6])
                ws.cell(row=i, column=7).value = toNumberFmt(ltTxn[6])
                allProdAmt = allProdAmt + toNumberFmt(ltTxn[7])
                ws.cell(row=i, column=8).value = toNumberFmt(ltTxn[7])
                allErrFee = allErrFee + toNumberFmt(ltTxn[8])
                ws.cell(row=i, column=9).value = toNumberFmt(ltTxn[8])
                allCost = allCost + toNumberFmt(ltTxn[5] + ltTxn[6] + ltTxn[7] + ltTxn[8])
                ws.cell(row=i, column=10).value = toNumberFmt(ltTxn[5] + ltTxn[6] + ltTxn[7] + ltTxn[8])
                allStlmAmt = allStlmAmt + toNumberFmt(ltTxn[4]) - toNumberFmt(ltTxn[5] + ltTxn[6] + ltTxn[7] + ltTxn[8])
                ws.cell(row=i, column=11).value = toNumberFmt(ltTxn[4]) - toNumberFmt(ltTxn[5] + ltTxn[6] + ltTxn[7] + ltTxn[8])
                i = i + 1
            # �ܼ�
            ws.cell(row=i, column=1).value = '�ܼ�:'
            ws.cell(row=i, column=5).value = allTxnAmt
            ws.cell(row=i, column=6).value = allIssAmt
            ws.cell(row=i, column=7).value = allSwtAmt
            ws.cell(row=i, column=8).value = allProdAmt
            ws.cell(row=i, column=9).value = allErrFee
            ws.cell(row=i, column=10).value = allCost
            ws.cell(row=i, column=11).value = allStlmAmt


    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_date)
    filename = filePath + 'AcqStlmCheckFile01_%s.xlsx' % stlm_date
    wb.save(filename)

    sql = "insert into TBL_STLM_TASK_CTL (host_date, " \
          "chnl_amt, " \
          "bal_mark) values ('%s', %.2f, '%s')" % (stlm_date, chnlbill.txnAmt, bal_sta)
    cursor = db.cursor()
    cursor.execute(sql)
    db.commit()
    cursor.close()

    db.close()

if __name__ == '__main__':
    main()