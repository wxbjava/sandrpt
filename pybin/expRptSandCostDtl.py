#!/home/acqbat/python36/bin/python3
#-*- coding:gb18030 -*-
#�����ɱ���ϸ��

import cx_Oracle
import sys
import os
from openpyxl.workbook import Workbook
from utl.common import *


def getStlmMd(stlmMd):
    if stlmMd == '0':
        return "S0"
    elif stlmMd == '1':
        return "T1"
    else:
        return "δ֪"

def getMchtType(mchtType):
    if mchtType == '1':
        return "��׼"
    elif mchtType == '2':
        return "�Ż�"
    else:
        return "����"

def getCardType(cardType):
    if cardType == '00':
        return "����"
    elif cardType == '01':
        return "���"
    else:
        return "δ֪"

def getAgentIncome(db, keyRsp, stlmDate):
    sql = "select PROFITS_AMT from TBL_INS_PROFITS_TXN_DTL where key_rsp = '%s' and host_date ='%s'" % (keyRsp, stlmDate)
    cursor = db.cursor()
    cursor.execute(sql)
    x = cursor.fetchone()
    if x is not None:
        return toNumberFmt(x[0])
    else:
        return 0


def newSandCostFileHead(ws):
    i = 1
    ws.cell(row=i, column=1).value = '��������'
    ws.cell(row=i, column=2).value = '����ʱ��'
    ws.cell(row=i, column=3).value = '�̻���'
    ws.cell(row=i, column=4).value = '������'
    ws.cell(row=i, column=5).value = '�̻�����'
    ws.cell(row=i, column=6).value = '��/��'
    ws.cell(row=i, column=7).value = '���׽��'
    ws.cell(row=i, column=8).value = '�̻�������'
    ws.cell(row=i, column=9).value = '�����з����'
    ws.cell(row=i, column=10).value = '��������ת�ӷ�'
    ws.cell(row=i, column=11).value = 'Ʒ�Ʒ����'
    ws.cell(row=i, column=12).value = '�̻������ʽ�'
    ws.cell(row=i, column=13).value = '�յ�����'
    ws.cell(row=i, column=14).value = '�ܲ�����'
    ws.cell(row=i, column=15).value = '��������'

def tailSandCostBody(ws,i,ltData, db, stlmDate):
    ws.cell(row=i, column=1).value = ltData[1]
    ws.cell(row=i, column=2).value = ltData[2]
    ws.cell(row=i, column=3).value = ltData[3]
    ws.cell(row=i, column=4).value = getStlmMd(ltData[4])
    ws.cell(row=i, column=5).value = getMchtType(ltData[5])
    ws.cell(row=i, column=6).value = getCardType(ltData[6])
    ws.cell(row=i, column=7).value = toNumberFmt(ltData[7])
    ws.cell(row=i, column=8).value = toNumberFmt(ltData[8])
    ws.cell(row=i, column=9).value = toNumberFmt(ltData[9])
    ws.cell(row=i, column=10).value = toNumberFmt(ltData[10])
    ws.cell(row=i, column=11).value = toNumberFmt(ltData[11])
    ws.cell(row=i, column=12).value = toNumberFmt(ltData[7] - ltData[8])
    ws.cell(row=i, column=13).value = toNumberFmt(ltData[8] - ltData[9] - ltData[10] - ltData[11])
    agentIncome = getAgentIncome(db, ltData[12], stlmDate)
    ws.cell(row=i, column=14).value = toNumberFmt(ltData[8] - ltData[9] - ltData[10] - ltData[11] - agentIncome)
    ws.cell(row=i, column=15).value = agentIncome

def main():
    db = cx_Oracle.connect('%s/%s@%s' % (os.environ['DBUSER'], os.environ['DBPWD'], os.environ['TNSNAME']),encoding='gb18030')
    if len(sys.argv) == 1:
        cursor = db.cursor()
        sql = "select BF_STLM_DATE from TBL_BAT_CUT_CTL"
        cursor.execute(sql)
        x = cursor.fetchone()
        stlm_date = x[0]
        cursor.close()
    else:
        stlm_date = sys.argv[1]

    print('hostDate %s expRptSandCost begin' % stlm_date)
    filePath = '%s/%s/' % (os.environ['RPT7HOME'], stlm_date)
    sql = "select trim(INS_ID_CD), txn_date, txn_time, MCHT_CD, STLM_MD, MCHT_TYPE, CARD_TYPE_2, " \
          "REAL_TRANS_AMT, mcht_fee, iss_fee, swt_fee, prod_fee, " \
          "KEY_RSP from tbl_stlm_txn_bill_dtl where host_date ='%s' and CHNL_ID='A001' order by INS_ID_CD,txn_date, txn_time" % stlm_date
    print(sql)
    cursor = db.cursor()
    cursor.execute(sql)
    insIdCdTmp = ''
    for ltData in cursor:
        if insIdCdTmp == '':
            insIdCdTmp = ltData[0]
            filename = filePath + 'Sand_Cost_Dtl_%s_%s.xlsx' % (ltData[0], stlm_date)
            wb = Workbook()
            ws = wb.active
            newSandCostFileHead(ws)
            i = 2
        if insIdCdTmp != ltData[0]:
            #�ر�ǰһ�������ļ�
            wb.save(filename)
            wb.close()
            #�´����������ļ�
            insIdCdTmp = ltData[0]
            filename = filePath + 'Sand_Cost_Dtl_%s_%s.xlsx' % (ltData[0], stlm_date)
            wb = Workbook()
            ws = wb.active
            newSandCostFileHead(ws)
            i = 2
        #д���ļ�
        tailSandCostBody(ws,i,ltData, db, stlm_date)
        i = i + 1
    if insIdCdTmp != '':
        wb.save(filename)
        wb.close()
    cursor.close()
    print('hostDate %s expRptSandCostDtl end' % stlm_date)


if __name__ == '__main__':
    main()